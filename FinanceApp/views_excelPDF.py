from io import BytesIO
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Q
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.template.loader import get_template

import tempfile
import os
import io
from xhtml2pdf import pisa 

from .models import ChartOfAccounts, GeneralLedger, JournalEntry, JournalLine

def get_account_balance(account, as_at_date=None):
    """Calculate net balance for a given account using GeneralLedger."""
    ledger = GeneralLedger.objects.filter(account=account).first()
    return ledger.current_balance if ledger else Decimal('0.00')

def get_accounts_by_type(account_type):
    """Return queryset of accounts filtered by type (ASSET, LIABILITY, etc.)."""
    return ChartOfAccounts.objects.filter(account_type=account_type, is_active=True)

from decimal import Decimal
from .models import GeneralLedger

def build_trial_balance():
    ledgers = GeneralLedger.objects.select_related('account').all()
    data = []
    for lg in ledgers:
        bal = lg.current_balance
        if bal > 0:
            data.append({
                'code': lg.account.accountno,
                'name': lg.account.name,
                'debit': bal,
                'credit': Decimal('0.00'),
            })
        else:
            data.append({
                'code': lg.account.accountno,
                'name': lg.account.name,
                'debit': Decimal('0.00'),
                'credit': -bal,
            })
    return data

def build_profit_loss(as_at_date=None):
    """Return income and expense accounts."""
    income = []
    expense = []
    for acc in get_accounts_by_type('INCOME'):
        income.append({'code': acc.accountno, 'name': acc.name, 'balance': get_account_balance(acc, as_at_date)})
    for acc in get_accounts_by_type('EXPENSE'):
        expense.append({'code': acc.accountno, 'name': acc.name, 'balance': get_account_balance(acc, as_at_date)})
    total_income = sum(i['balance'] for i in income)
    total_expense = sum(e['balance'] for e in expense)
    net_profit = total_income - total_expense
    return {
        'income': income, 'total_income': total_income,
        'expense': expense, 'total_expense': total_expense,
        'net_profit': net_profit,
    }

# ----- Views -----
@login_required
def trial_balance_report(request, format='html'):
    """Trial Balance report in HTML, PDF, or Excel."""
    data = build_trial_balance()
    context = {'data': data, 'title': 'Trial Balance', 'date': request.GET.get('date')}
    if format == 'pdf':
        return render_to_pdf('FinanceApp/trial_balance.html', context)
    elif format == 'excel':
        return export_to_excel(data, 'Trial Balance')
    return render(request, 'FinanceApp/trial_balance.html', context)

@login_required
def balance_sheet_report(request, format='html'):
    data = build_balance_sheet()
    context = {'data': data, 'title': 'Balance Sheet', 'date': request.GET.get('date')}
    if format == 'pdf':
        return render_to_pdf('FinanceApp/balance_sheet.html', context)
    elif format == 'excel':
        return export_balance_sheet_excel(data, 'Balance Sheet')
    return render(request, 'FinanceApp/balance_sheet.html', context)

@login_required
def profit_loss_report(request, format='html'):
    data = build_profit_loss()
    context = {'data': data, 'title': 'Profit & Loss Statement', 'date': request.GET.get('date')}
    if format == 'pdf':
        return render_to_pdf('FinanceApp/profit_loss.html', context)
    elif format == 'excel':
        return export_profit_loss_excel(data, 'Profit & Loss')
    return render(request, 'FinanceApp/profit_loss.html', context)

@login_required
def ledger_balances_report(request, format='html'):
    """All ledger accounts with current balance."""
    ledgers = GeneralLedger.objects.select_related('account').filter(account__is_active=True).order_by('account__accountno')
    data = [{'code': l.account.accountno, 'name': l.account.name, 'balance': l.current_balance} for l in ledgers]
    context = {'data': data, 'title': 'Ledger Balances', 'date': request.GET.get('date')}
    if format == 'pdf':
        return render_to_pdf('FinanceApp/ledger_balances.html', context)
    elif format == 'excel':
        return export_ledger_excel(data, 'Ledger Balances')
    return render(request, 'FinanceApp/ledger_balances.html', context)


from decimal import Decimal
from django.utils import timezone
from .models import ChartOfAccounts, GeneralLedger

def get_trial_balance_data(as_at_date=None):
    """Return trial balance data with totals and balance check."""
    # If no date, use today
    if as_at_date is None:
        as_at_date = timezone.now().date()
    
    accounts = ChartOfAccounts.objects.filter(is_active=True).order_by('accountno')
    trial_balance = []
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    
    for account in accounts:
        # Get current balance from GeneralLedger (you already have this)
        ledger = GeneralLedger.objects.filter(account=account).first()
        balance = ledger.current_balance if ledger else Decimal('0.00')
        
        if balance > 0:
            debit = balance
            credit = Decimal('0.00')
        else:
            debit = Decimal('0.00')
            credit = -balance
        
        trial_balance.append({
            'account': account,
            'debit': debit,
            'credit': credit,
        })
        total_debit += debit
        total_credit += credit
    
    difference = total_debit - total_credit
    is_balanced = abs(difference) < Decimal('0.01')
    
    return {
        'trial_balance': trial_balance,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'difference': difference,
        'is_balanced': is_balanced,
        'as_at_date': as_at_date,
    }

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io


def test_pdf(request):
    """Simplest possible PDF generation test."""
    from io import BytesIO
    from xhtml2pdf import pisa
    from django.http import HttpResponse

    html = "<html><body><h1>PDF Test</h1></body></html>"
    result = BytesIO()
    pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    # Ensure we return an HttpResponse
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="test.pdf"'
    return response


def trial_balance_pdf(request, format=None):   # <-- add format parameter
    try:
        # Get date
        date_str = request.GET.get('as_at_date')
        if date_str:
            from datetime import datetime
            as_at_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            as_at_date = timezone.now().date()

        # Get data
        context = get_trial_balance_data(as_at_date)

        # Render template
        template = get_template('FinanceApp/trial_balance.html')
        html = template.render(context)

        # Generate PDF
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

        if pdf is None or pdf.err:
            return HttpResponse("PDF generation failed", status=500)

        # ✅ Return proper HttpResponse
        response = HttpResponse(
            result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="trial_balance.pdf"'
        return response

    except Exception as e:
        # ✅ Always return HttpResponse on error
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)
    
    
def trial_balance_list(request, format=None):   # <-- add format parameter
    try:
        # Get date
        date_str = request.GET.get('as_at_date')
        if date_str:
            from datetime import datetime
            as_at_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            as_at_date = timezone.now().date()

        # Get data
        context = get_trial_balance_data(as_at_date)

        # Render template
        template = get_template('FinanceApp/trial_balance_list.html')
        html = template.render(context)

        # Generate PDF
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

        if pdf is None or pdf.err:
            return HttpResponse("PDF generation failed", status=500)

        # ✅ Return proper HttpResponse
        response = HttpResponse(
            result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="trial_balance.pdf"'
        return response

    except Exception as e:
        # ✅ Always return HttpResponse on error
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


# FinanceApp/views.py

def trial_balance(request):
    """Display the Trial Balance as an HTML page (printable)."""
    date_str = request.GET.get('as_at_date')
    if date_str:
        from datetime import datetime
        as_at_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        as_at_date = timezone.now().date()

    context = get_trial_balance_data(as_at_date)
    return render(request, 'FinanceApp/trial_balance.html', context)


def trial_balance_excel(request):
    date_str = request.GET.get('as_at_date')
    if date_str:
        from datetime import datetime
        as_at_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    else:
        as_at_date = timezone.now().date()
    
    data = get_trial_balance_data(as_at_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.append(['Account Code', 'Account Name', 'Debit (₵)', 'Credit (₵)'])
    for item in data['trial_balance']:
        ws.append([item['account'].accountno, item['account'].name, float(item['debit']), float(item['credit'])])
    ws.append(['', 'TOTAL', float(data['total_debit']), float(data['total_credit'])])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="trial_balance.xlsx"'
    wb.save(response)
    return response

# ----- Helpers for PDF and Excel -----
from django.template.loader import get_template
from django.http import HttpResponse
import io
from xhtml2pdf import pisa   # <-- correct import




def export_to_excel(data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Headers
    headers = ['Account Code', 'Account Name', 'Debit (₵)', 'Credit (₵)']
    ws.append(headers)
    for row in data:
        ws.append([row['code'], row['name'], float(row['debit']), float(row['credit'])])
    # Style
    for col in ['A','B','C','D']:
        ws.column_dimensions[col].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_name}.xlsx"'
    wb.save(response)
    return response

def export_balance_sheet_excel(data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['ASSETS', '', 'Amount (₵)'])
    for a in data['assets']:
        ws.append([a['name'], a['code'], float(a['balance'])])
    ws.append(['Total Assets', '', float(data['total_assets'])])
    ws.append([])
    ws.append(['LIABILITIES', '', 'Amount (₵)'])
    for l in data['liabilities']:
        ws.append([l['name'], l['code'], float(l['balance'])])
    ws.append(['Total Liabilities', '', float(data['total_liabilities'])])
    ws.append([])
    ws.append(['EQUITY', '', 'Amount (₵)'])
    for e in data['equity']:
        ws.append([e['name'], e['code'], float(e['balance'])])
    ws.append(['Total Equity', '', float(data['total_equity'])])
    ws.append([])
    ws.append(['TOTAL LIABILITIES + EQUITY', '', float(data['total_liabilities'] + data['total_equity'])])
    for col in ['A','B','C']:
        ws.column_dimensions[col].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_name}.xlsx"'
    wb.save(response)
    return response

def export_profit_loss_excel(data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['INCOME', '', 'Amount (₵)'])
    for i in data['income']:
        ws.append([i['name'], i['code'], float(i['balance'])])
    ws.append(['Total Income', '', float(data['total_income'])])
    ws.append([])
    ws.append(['EXPENSES', '', 'Amount (₵)'])
    for e in data['expense']:
        ws.append([e['name'], e['code'], float(e['balance'])])
    ws.append(['Total Expenses', '', float(data['total_expense'])])
    ws.append([])
    ws.append(['NET PROFIT/(LOSS)', '', float(data['net_profit'])])
    for col in ['A','B','C']:
        ws.column_dimensions[col].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_name}.xlsx"'
    wb.save(response)
    return response

def export_ledger_excel(data, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(['Account Code', 'Account Name', 'Balance (₵)'])
    for row in data:
        ws.append([row['code'], row['name'], float(row['balance'])])
    for col in ['A','B','C']:
        ws.column_dimensions[col].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_name}.xlsx"'
    wb.save(response)
    return 


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io
from openpyxl import Workbook
from .models import GeneralLedger

def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return None
    return result.getvalue()

def ledger_balances_pdf(request):
    ledgers = GeneralLedger.objects.select_related('account').all()
    context = {
        'ledgers': ledgers,
        'title': 'Ledger Balances',
        'date': request.GET.get('date', 'Today'),
    }
    pdf_bytes = render_to_pdf('FinanceApp/ledger_balances.html', context)
    if pdf_bytes is None:
        return HttpResponse("Error generating PDF", status=500)
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ledger_balances.pdf"'
    return response

def ledger_balances_excel(request):
    ledgers = GeneralLedger.objects.select_related('account').all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger Balances"
    ws.append(['Account Code', 'Account Name', 'Balance (₵)'])
    for lg in ledgers:
        ws.append([lg.account.accountno, lg.account.name, float(lg.current_balance)])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ledger_balances.xlsx"'
    wb.save(response)
    return response

from openpyxl import Workbook
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io

def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    if pdf.err:
        return None
    return result.getvalue()

def ledger_statement_pdf(request):
    # Same queryset as the list view (apply same filters)
    ledgers = GeneralLedger.objects.select_related('account').all().order_by('account__accountno')
    # apply filters (duplicate logic; better to extract a helper)
    # ... (copy filtering from ledger_statement)
    context = {'ledgers': ledgers, 'date': timezone.now()}
    pdf = render_to_pdf('FinanceApp/ledger_statement_pdf.html', context)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ledger_statement.pdf"'
    return response

def ledger_statement_excel(request):
    ledgers = GeneralLedger.objects.select_related('account').all().order_by('account__accountno')
    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger Statement"
    ws.append(['Account No', 'Account Name', 'Account Type', 'Current Balance (₵)'])
    for lg in ledgers:
        ws.append([lg.account.accountno, lg.account.name, lg.account.account_type, float(lg.current_balance)])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ledger_statement.xlsx"'
    wb.save(response)
    return response