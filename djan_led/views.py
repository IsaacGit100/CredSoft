from django.shortcuts import render

# Create your views here.


def djan_led_home(request):
    return render(request, 'djan_led/djan_led_home.html')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

# from django_ledger.io.cursor import JournalEntryCursor   # <-- Correct import
from django.db.models import Sum
from decimal import Decimal
from django.utils import timezone
from .models import UserProfile
from .utils import user_can_access_entity
from django_ledger.models import (
    EntityModel,
    JournalEntryModel,
    TransactionModel,
    AccountModel,
    LedgerModel,
)

@login_required
def coa_home(request, slug):
    return render(request, 'djan_led/coa_home.html')

@login_required
def supervisor_trans_home(request, slug):
    return render(request, 'djan_led/supervisor_trans_home.html')

@login_required
def supervisor_cred_home(request, slug):
    return render(request, 'djan_led/supervisor_creditunion_home.html')


@login_required
def after_login_redirect(request):
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    # ---- POS User: go straight to POS dashboard ----
    if profile.role == "pos":
        if profile.default_entity:
            return redirect("pos:sales_form", slug=profile.default_entity.slug)
        else:
            return render(
                request,
                "djan_led/no_entity.html",
                {"message": "No entity assigned for POS."},
            )
            
    if profile.role == 'technical':
        return redirect('Tech:tech_dashboard')
    
    if profile.role == 'Church':
        return redirect('ChurchApp:church_dashboard', slug=profile.default_entity.slug)
    
        



    if profile.role == "CreditUnion":
        if profile.default_entity:
            return redirect("CreditUnion:union_dashboard", slug=profile.default_entity.slug)
        else:
            return render(
                request,
                "djan_led/no_entity.html",
                {"message": "No entity assigned for POS."},
            )

    # ---- Super Admin: portal ----
    if profile.role == "super_admin":
        return redirect("Consolidated:portal")

    # ---- Normal User: entity dashboard ----
    if profile.default_entity:
        return redirect("djan_led:entity_dashboard", slug=profile.default_entity.slug)

    entities = profile.allowed_entities.all()
    if entities.count() == 1:
        return redirect("djan_led:entity_dashboard", slug=entities.first().slug)
    elif entities.count() > 1:
        return render(request, "djan_led/select_entity.html", {"entities": entities})
    else:
        return render(
            request, "djan_led/no_entity.html", {"message": "No entity assigned."}
        )


@login_required
def entity_dashboard(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    name_lower = entity.name.lower()
    
     # Super admin or allowed user?
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    
    
    module = request.GET.get('module')
    if not module:
        if 'school' in entity.name.lower():
            module = 'school'
        elif 'credit union' in entity.name.lower():
            module = 'credit_union'
        elif 'church' in entity.name.lower():
            module = 'church'
        else:
            module = 'finance'
    
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    # Check access
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Get recent journal entries
    recent_entries = JournalEntryModel.objects.filter(
        ledger__entity=entity
    ).order_by('-created')[:10]

    # Cash balance (Account code 1010)
    try:
        cash_account = AccountModel.objects.get(
            coa_model__entity=entity,  # <-- Fixed
            code='1010'
        )
        cash_debits = TransactionModel.objects.filter(
            account=cash_account,
            tx_type='debit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        cash_credits = TransactionModel.objects.filter(
            account=cash_account,
            tx_type='credit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        cash_balance = cash_debits - cash_credits
    except AccountModel.DoesNotExist:
        cash_balance = Decimal('0')

    # Revenue total (Account 4010)
    try:
        revenue_account = AccountModel.objects.get(
            coa_model__entity=entity,  # <-- Fixed
            code='4010'
        )
        revenue_total = TransactionModel.objects.filter(
            account=revenue_account,
            tx_type='credit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    except AccountModel.DoesNotExist:
        revenue_total = Decimal('0')

    # Expense total (all expense accounts)
    expense_accounts = AccountModel.objects.filter(
        coa_model__entity=entity,  # <-- Fixed
        role='expense'
    )
    expense_total = TransactionModel.objects.filter(
        account__in=expense_accounts,
        tx_type='debit'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    net_income = revenue_total - expense_total

    context = {
        'entity': entity,
        'module': module,
        'recent_entries': recent_entries,
        'cash_balance': cash_balance,
        'revenue_total': revenue_total,
        'expense_total': expense_total,
        'net_income': net_income,
    }
    return render(request, 'djan_led/entity_dashboard.html', context)


@login_required
def chart_of_accounts(request, slug):
    
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    if coa:
        accounts = AccountModel.objects.filter(coa_model=coa).order_by('code')
    else:
        accounts = []
    return render(request, 'djan_led/chart_of_accounts.html', {'entity': entity, 'accounts': accounts})


# djan_led/views.py (add these functions after the existing ones)

@login_required
def journal_entries(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    """List all journal entries for an entity."""
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Check access (simplified – we'll assume user has permission)
    # entries = JournalEntryModel.objects.filter(entity=entity).order_by('-date')
    entries = JournalEntryModel.objects.filter(ledger__entity=entity).order_by('-created')
    return render(request, 'djan_led/journal_entries.html', {'entity': entity, 'entries': entries})

@login_required
def trial_balance(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)

    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Get all accounts for this entity (non‑root, active)
    accounts = AccountModel.objects.filter(
        coa_model__entity=entity,
        active=True,
        depth__gt=1
    ).order_by('code')

    trial_data = []
    total_debits = Decimal('0')
    total_credits = Decimal('0')

    for acc in accounts:
        debit_total = TransactionModel.objects.filter(
            account=acc, tx_type='debit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        credit_total = TransactionModel.objects.filter(
            account=acc, tx_type='credit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        # Skip accounts with zero balance
        if debit_total == 0 and credit_total == 0:
            continue

        balance = debit_total - credit_total
        total_debits += debit_total
        total_credits += credit_total

        trial_data.append({
            'code': acc.code,
            'name': acc.name,
            'debit': debit_total,
            'credit': credit_total,
            'balance': balance,
        })

    context = {
        'entity': entity,
        'trial_data': trial_data,
        'total_debits': total_debits,
        'total_credits': total_credits,
    }
    return render(request, 'djan_led/trial_balance.html', context)


@login_required
def trial_balance_all_accounts(request, slug):
    
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Check access
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Get all accounts – corrected to use coa_model__entity
    accounts = AccountModel.objects.filter(coa_model__entity=entity).order_by('code')

    trial_data = []
    for acc in accounts:
        debit_total = TransactionModel.objects.filter(
            account=acc, tx_type='debit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        credit_total = TransactionModel.objects.filter(
            account=acc, tx_type='credit'
        ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        balance = debit_total - credit_total

        trial_data.append({
            'code': acc.code,
            'name': acc.name,
            'debit': debit_total,
            'credit': credit_total,
            'balance': balance,
        })

    context = {
        'entity': entity,
        'trial_data': trial_data,
    }
    return render(request, 'djan_led/trial_balance.html', context)


@login_required
def income_statement(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Check access
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Revenue accounts (role='revenue')
    revenue_accounts = AccountModel.objects.filter(
        coa_model__entity=entity,  # <-- Fixed
        role='revenue'
    )
    revenue_total = TransactionModel.objects.filter(
        account__in=revenue_accounts,
        tx_type='credit'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    # Expense accounts (role='expense')
    expense_accounts = AccountModel.objects.filter(
        coa_model__entity=entity,  # <-- Fixed
        role='expense'
    )
    expense_total = TransactionModel.objects.filter(
        account__in=expense_accounts,
        tx_type='debit'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    net_income = revenue_total - expense_total

    context = {
        'entity': entity,
        'revenue_total': revenue_total,
        'expense_total': expense_total,
        'net_income': net_income,
    }
    return render(request, 'djan_led/income_statement.html', context)


@login_required
def journal_entry_detail(request, slug, pk):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    #entry = get_object_or_404(JournalEntryModel, entity=entity, pk=pk)
    entry = get_object_or_404(JournalEntryModel, ledger__entity=entity, pk=pk)
    transactions = TransactionModel.objects.filter(journal_entry=entry)
    return render(request, 'djan_led/journal_entry_detail.html', {'entity': entity, 'entry': entry, 'transactions': transactions})


from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
# from django_ledger.models import EntityModel
from .forms import CreateParishForm

@staff_member_required  # Only staff/admin can access
def create_parish(request):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    if request.method == 'POST':
        form = CreateParishForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            slug = form.cleaned_data['slug']
            parent = form.cleaned_data.get('parent_entity')

            # Create the entity
            if parent:
                # Create as child of parent
                entity = EntityModel.add_child(parent=parent, name=name, slug=slug, admin=request.user)
            else:
                # Create as root entity
                entity = EntityModel.add_root(name=name, slug=slug, admin=request.user,)

            # Create the Chart of Accounts
            coa = entity.create_chart_of_accounts(assign_as_default=True, commit=True, coa_name='Default COA')

            # Optional: Assign this user to the new entity
            try:
                profile = request.user.djan_led_profile
                profile.allowed_entities.add(entity)
                if not profile.default_entity:
                    profile.default_entity = entity
                    profile.save()
            except:
                pass

            messages.success(
                request,
                f" Parish '{name}' created successfully with Chart of Accounts!"
            )
            return redirect('entity_dashboard', slug=entity.slug)

    else:
        form = CreateParishForm()

#    account_type = request.POST.get("entity_type")  # e.g., 'church', 'school', etc.
#    if account_type:
#        root_nodes = (
#            root_assets,
#            root_liabilities,
#            root_capital,
#            root_income,
#            root_expenses,
#        )
#        accounts_list = get_accounts_for_type(account_type, root_nodes)
#        add_accounts_to_coa(coa, accounts_list)

    return render(request, 'djan_led/create_parish.html', {'form': form})


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
# from django_ledger.models import EntityModel, LedgerModel, JournalEntryModel, AccountModel, TransactionModel
from django.utils import timezone
from decimal import Decimal
from .forms import RecordOfferingForm

@login_required
def record_offering(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if request.method == 'POST':
        form = RecordOfferingForm(request.POST)
        if form.is_valid():
            date = form.cleaned_data['date']
            amount = form.cleaned_data['amount']
            description = form.cleaned_data['description'] or f'Offering - {entity.name}'
            
            # Get or create a ledger
            ledger, _ = LedgerModel.objects.get_or_create(
                entity=entity,
                name='Default Ledger',
                slug='default',
                defaults={'created_by': request.user}
            )
            
            # Create Journal Entry
            je = JournalEntryModel.objects.create(entity=entity, ledger=ledger, date=date, description=description, posted=True)
            
            # Get accounts (Cash and Revenue)
            cash = AccountModel.objects.get(coa_model__entity=entity, code='1010')
            revenue = AccountModel.objects.get(coa_model__entity=entity, code='4010')
            
            # Create transactions
            TransactionModel.objects.create(
                journal_entry=je,
                account=cash,
                amount=amount,
                tx_type='debit'
            )
            TransactionModel.objects.create(
                journal_entry=je,
                account=revenue,
                amount=amount,
                tx_type='credit'
            )
            
            messages.success(request, f" Offering of {amount} recorded successfully!")
            return redirect('djan_led:entity_dashboard', slug=entity.slug)
    else:
        form = RecordOfferingForm(initial={'date': timezone.now().date()})
    
    return render(request, 'djan_led/record_offering.html', {'form': form, 'entity': entity})

from django.contrib import messages

@login_required
def chart_of_accounts_create(request, slug):
  
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Check if user has access
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Get or create default CoA
    try:
        coa = entity.get_default_coa()
    except:
        coa = None
    
    if coa is None:
        # Create a default CoA
        coa = entity.create_chart_of_accounts(
            assign_as_default=True,
            commit=True,
            coa_name='Default COA'
        )
        messages.info(request, f"Chart of Accounts created for {entity.name}")
    
    accounts = AccountModel.objects.filter(coa_model=coa).order_by('code')
    
    context = {
        'entity': entity,
        'accounts': accounts,
    }
    return render(request, 'djan_led/chart_of_accounts.html', context)

from django.contrib import messages

@login_required
def chart_of_accounts(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Check if user has access
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    
    # if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
    #    return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Get or create default CoA
    try:
        coa = entity.get_default_coa()
    except:
        coa = None
    
    if coa is None:
        # Create a default CoA
        coa = entity.create_chart_of_accounts(
            assign_as_default=True,
            commit=True,
            coa_name='Default COA'
        )
        messages.info(request, f"Chart of Accounts created for {entity.name}")
    
    accounts = AccountModel.objects.filter(coa_model=coa).order_by('code')
    
    context = {
        'entity': entity,
        'accounts': accounts,
    }
    return render(request, 'djan_led/chart_of_accounts.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages

from .forms import AddAccountForm

# djan_led/views.# djan_led/views.py
@login_required

def add_account(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    from django_ledger.models import AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    #    User can access all
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    ##
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found. Please autofill first.")
        return redirect('chart_of_accounts', slug=entity.slug)

    if request.method == 'POST':
        form = AddAccountForm(request.POST, entity=entity)  # Pass entity
        if form.is_valid():
            parent = form.cleaned_data['parent']
            code = form.cleaned_data['code']
            name = form.cleaned_data['name']
            role = form.cleaned_data['role']
            balance_type = form.cleaned_data['balance_type']

            # Check if account code already exists in this COA
            if AccountModel.objects.filter(coa_model=coa, code=code).exists():
                messages.error(request, f"Account with code {code} already exists.")
                return render(request, 'djan_led/add_account.html', {'form': form, 'entity': entity})

            # Create the account
            acc = AccountModel.add_root(coa_model=coa, code=code, name=name, role=role, balance_type=balance_type)
            acc.move(parent, pos='last-child')

            messages.success(request, f"Account {code} - {name} added successfully.")
            return redirect('djan_led:chart_of_accounts', slug=entity.slug)
    else:
        form = AddAccountForm(entity=entity)  # Pass entity

    return render(request, 'djan_led/add_account.html', {'form': form, 'entity': entity})


@login_required
def balance_sheet(request, slug):
    from django_ledger.models import (
        EntityModel,
        JournalEntryModel,
        TransactionModel,
        AccountModel,
    )
    from decimal import Decimal
    from django.db.models import Sum

    entity = get_object_or_404(EntityModel, slug=slug)

    if not user_can_access_entity(request.user, entity):
        return render(request, "djan_led/access_denied.html", {"entity": entity})

    coa = entity.get_default_coa()
    if not coa:
        return render(
            request,
            "djan_led/balance_sheet.html",
            {"entity": entity, "error": "No COA"},
        )

    # Helper to compute account balance
    def get_balance(acc):
        debits = TransactionModel.objects.filter(
            account=acc, tx_type="debit"
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0")
        credits = TransactionModel.objects.filter(
            account=acc, tx_type="credit"
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0")
        if acc.balance_type == "debit":
            return debits - credits
        else:
            return credits - debits

    # Get active, non‑root accounts for each category
    accounts = AccountModel.objects.filter(
        coa_model=coa, active=True, depth__gt=1  # exclude root nodes
    )

    asset_data = []
    liability_data = []
    equity_data = []

    for acc in accounts:
        balance = get_balance(acc)
        if balance == 0:
            continue  # skip zero‑balance accounts

        item = {"code": acc.code, "name": acc.name, "balance": balance}
        if acc.role == "asset":
            asset_data.append(item)
        elif acc.role == "liability":
            liability_data.append(item)
        elif acc.role == "equity":
            equity_data.append(item)
        # revenue/expense accounts are not shown on balance sheet

    total_assets = sum(item["balance"] for item in asset_data)
    total_liabilities = sum(item["balance"] for item in liability_data)
    total_equity = sum(item["balance"] for item in equity_data)

    context = {
        "entity": entity,
        "asset_data": asset_data,
        "liability_data": liability_data,
        "equity_data": equity_data,
        "total_assets": total_assets,
        "total_liabilities": total_liabilities,
        "total_equity": total_equity,
    }
    return render(request, "djan_led/balance_sheet.html", context)


@login_required
def balance_sheet_all_accounts(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    from django_ledger.models import AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    # Check access (same as other views)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    coa = entity.get_default_coa()
    if not coa:
        return render(request, 'djan_led/balance_sheet.html', {'entity': entity, 'error': 'No COA'})

    # Get all accounts grouped by role
    accounts = AccountModel.objects.filter(coa_model=coa)
    assets = accounts.filter(role='asset')
    liabilities = accounts.filter(role='liability')
    equity = accounts.filter(role='equity')

    # Calculate balances for each account
    def get_balance(acc):
        debits = TransactionModel.objects.filter(account=acc, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        credits = TransactionModel.objects.filter(account=acc, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        if acc.balance_type == 'debit':
            return debits - credits
        else:
            return credits - debits

    asset_data = [{'code': a.code, 'name': a.name, 'balance': get_balance(a)} for a in assets]
    liability_data = [{'code': l.code, 'name': l.name, 'balance': get_balance(l)} for l in liabilities]
    equity_data = [{'code': e.code, 'name': e.name, 'balance': get_balance(e)} for e in equity]

    total_assets = sum(item['balance'] for item in asset_data)
    total_liabilities = sum(item['balance'] for item in liability_data)
    total_equity = sum(item['balance'] for item in equity_data)

    context = {
        'entity': entity,
        'asset_data': asset_data,
        'liability_data': liability_data,
        'equity_data': equity_data,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
    }
    return render(request, 'djan_led/balance_sheet.html', context)

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required


@login_required
def autofill_chart_of_accounts(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    coa = entity.get_default_coa()

    if not coa:
        # Create a default COA if missing
        coa = entity.create_chart_of_accounts(
            assign_as_default=True,
            commit=True,
            coa_name='Default COA'
        )
        messages.info(request, f"Created default Chart of Accounts for {entity.name}.")

    # ----- Define the accounts -----
    accounts_data = [
        # Root nodes
        ("1000", "ASSETS", "asset", "debit", None),
        ("2000", "LIABILITIES", "liability", "credit", None),
        ("3000", "EQUITY", "equity", "credit", None),
        ("4000", "REVENUE", "revenue", "credit", None),
        ("5000", "EXPENSES", "expense", "debit", None),
        # Assets sub‑nodes
        ("1100", "Non-Current Assets", "asset", "debit", "1000"),
        ("1110", "Property, Plant & Equipment", "asset", "debit", "1100"),
        ("1120", "Investments", "asset", "debit", "1100"),
        #
        ("1200", "Current Assets", "asset", "debit", "1000"),
        ("1210", "Cash", "asset", "debit", "1200"),
        ("1220", "Bank Accounts", "asset", "debit", "1200"),
        ("1230", "Accounts Receivable", "asset", "debit", "1200"),
        # Asset details
        ("1111", "Land", "asset", "debit", "1110"),
        ("1112", "Buildings", "asset", "debit", "1110"),
        ("1113", "Vehicles", "asset", "debit", "1110"),
        ("1114", "Furniture & Equipment", "asset", "debit", "1110"),
        
        ("1121", "Treasury Bills", "asset", "debit", "1120"),
        ("1211", "Cash - Main", "asset", "debit", "1210"),
        ("1212", "Cash - Petty", "asset", "debit", "1210"),
        ("1221", "Bank - General", "asset", "debit", "1220"),
        ("1222", "Bank - Building Fund", "asset", "debit", "1220"),
        ("1223", "Bank - Mission Fund", "asset", "debit", "1220"),
        
        ("1231", "Debtors", "asset", "debit", "1230"),
        ("1232", "Other Receivables", "asset", "debit", "1230"),
        # Liabilities
        ("2100", "Current Liabilities", "liability", "credit", "2000"),
        ("2110", "Accounts Payable", "liability", "credit", "2100"),
        ("2200", "Non-Current Liabilities", "liability", "credit", "2000"),
        ("2111", "Creditors", "liability", "credit", "2110"),
        ("2112", "Accrued Expenses", "liability", "credit", "2110"),
        ("2210", "Loans Payable", "liability", "credit", "2200"),
        # Equity
        ("3010", "Owner's Equity", "equity", "credit", "3000"),
        ("3020", "Retained Earnings", "equity", "credit", "3000"),
        # Revenue
        ("4010", "General Offerings", "revenue", "credit", "4000"),
        #
        ("4011", "Day Borns", "revenue", "credit", "4000"),
        ("4011.01", "Day Borns - Sunday", "revenue", "credit", "4000"),
        ("4011.02", "Day Borns - Monday", "revenue", "credit", "4000"),
        ("4011.03", "Day Borns - Tuesday", "revenue", "credit", "4000"),
        ("4011.04", "Day Borns - Wednesday", "revenue", "credit", "4000"),
        ("4011.05", "Day Borns - Thursday", "revenue", "credit", "4000"),
        ("4011.06", "Day Borns - Friday", "revenue", "credit", "4000"),
        ("4011.07", "Day Borns - Saturday", "revenue", "credit", "4000"),
        #
        ("4012", "Guilds", "revenue", "credit", "4000"),
        ("4012.01", "Guilds - Servers", "revenue", "credit", "4000"),
        ("4012.02", "Guilds - Shepherd", "revenue", "credit", "4000"),
        ("4012.03", "Guilds - Women Felloship", "revenue", "credit", "4000"),
        ("4012.04", "Guilds - Men Fellowship", "revenue", "credit", "4000"),
        ("4012.05", "Guilds - Mothers Union", "revenue", "credit", "4000"),
        ("4012.06", "Guilds - Church Choir", "revenue", "credit", "4000"),
        ("4012.07", "Guilds - Club", "revenue", "credit", "4000"),
        ("4012.08", "Guilds - BrotherHood", "revenue", "credit", "4000"),
        ("4012.09", "Guilds - St Francis Guild", "revenue", "credit", "4000"),
        ("4012.10", "Guilds - AYPA", "revenue", "credit", "4000"),
        #
        ("4013", "Harvest Thank Offering", "revenue", "credit", "4000"),
        ("4013.01", "Harvest - Children", "revenue", "credit", "4000"),
        ("4013.02", "Harvest - Men", "revenue", "credit", "4000"),
        ("4013.03", "Harvest - Women", "revenue", "credit", "4000"),
        ("4013.04", "Harvest - Guilds", "revenue", "credit", "4000"),
        ("4013.05", "Harvest - Main", "revenue", "credit", "4000"),
        #
        ("4014", "Dues", "revenue", "credit", "4000"),
        ("4014.01", "Member Dues", "revenue", "credit", "4000"),
        #
        ("4015", "Tithes", "revenue", "credit", "4000"),
        ("4015.01", "Tithes", "revenue", "credit", "4000"),
        #
        ("4016", "Thank Offering", "revenue", "credit", "4000"),
        ("4016.01", "Special Thank Offering", "revenue", "credit", "4000"),
        #
        #
        ("4017", "Easter Offerings", "revenue", "credit", "4000"),
        ("4017.01", "Easter - Self Denial", "revenue", "credit", "4000"),
        ("4017.02", "Easter - Stations Of The Cross", "revenue", "credit", "4000"),
        ("4017.03", "Easter - Maudy Thursday", "revenue", "credit", "4000"),
        ("4017.04", "Easter - Veneration Of The Cross", "revenue", "credit", "4000"),
        ("4017.05", "Easter - 3-Hours Service", "revenue", "credit", "4000"),
        ("4017.06", "Easter - Holy Saturday Rites", "revenue", "credit", "4000"),
        ("4017.07", "Easter - Easter Day", "revenue", "credit", "4000"),
        #
        ("4018", "Christmas", "revenue", "credit", "4000"),
        ("4018.01", "Christmas Offering", "revenue", "credit", "4000"),
        
        ("4020", "Donations", "revenue", "credit", "4000"),
        
        ("4030", "Special Events", "revenue", "credit", "4000"),
        ("4040", "Rental Income", "revenue", "credit", "4000"),
        # Expenses
        ("5010", "Salaries & Wages", "expense", "debit", "5000"),
        ("5010.01", "Salaries & Wages", "expense", "debit", "5000"),
        #
        ("5020", "Rent", "expense", "debit", "5000"),
        ("5030", "Utilities", "expense", "debit", "5000"),
        ("5030.01","Utilities - Ghana Water", "expense", "debit", "5000"),
        ("5030.02", "Utilities - ECG", "expense", "debit", "5000"),
        ("5030.03", "Utilities - Comms", "expense", "debit", "5000"),
        
        ("5040", "Office Supplies", "expense", "debit", "5000"),
        ("5050", "Maintenance", "expense", "debit", "5000"),
        ("5060", "Insurance", "expense", "debit", "5000"),
        ("5070", "Transport", "expense", "debit", "5000"),
    ]

    created_count = 0
    skipped_count = 0

    for code, name, role, balance_type, parent_code in accounts_data:
        # Skip if account already exists in this COA
        if AccountModel.objects.filter(coa_model=coa, code=code).exists():
            skipped_count += 1
            continue

        # Create the account
        acc = AccountModel.add_root(coa_model=coa, code=code, name=name, role=role, balance_type=balance_type)

        # Move under parent if specified
        if parent_code:
            try:
                parent = AccountModel.objects.get(coa_model=coa, code=parent_code)
                acc.move(parent, pos='last-child')
            except AccountModel.DoesNotExist:
                # Parent not found – leave it at root
                pass

        created_count += 1

    messages.success(
        request,
        f" Autofill complete! {created_count} new accounts added, {skipped_count} already existed."
    )

    return redirect('djan_led:chart_of_accounts', slug=entity.slug)


from django.db.models import Sum, Q
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404


@login_required
def cash_flow(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    # Access check
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Get all cash accounts (Cash and Bank) – we use code prefixes
    # Cash: codes starting with '1210' (like 1211, 1212, etc.)
    # Bank: codes starting with '1220' (like 1221, 1222, etc.)
    cash_codes_start = ['1210', '1220']
    cash_accounts = AccountModel.objects.filter(
        coa_model__entity=entity,
        role='asset',
        balance_type='debit'
    ).filter(
        Q(code__startswith='1210') | Q(code__startswith='1220')
    )
    cash_uuids = [acc.uuid for acc in cash_accounts]

    # Cash inflows: debits to cash accounts
    cash_in = TransactionModel.objects.filter(
        account__uuid__in=cash_uuids,
        tx_type='debit'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    # Cash outflows: credits to cash accounts
    cash_out = TransactionModel.objects.filter(
        account__uuid__in=cash_uuids,
        tx_type='credit'
    ).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

    net_cash = cash_in - cash_out

    context = {
        'entity': entity,
        'cash_in': cash_in,
        'cash_out': cash_out,
        'net_cash': net_cash,
    }
    return render(request, 'djan_led/cash_flow.html', context)

from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404


@login_required
def chart_of_accounts_pdf(request, slug):
    from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
    entity = get_object_or_404(EntityModel, slug=slug)
    # Access check (optional but good practice)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    coa = entity.get_default_coa()
    if not coa:
        return HttpResponse("No Chart of Accounts found for this entity.", status=404)

    # Order by treebeard's path for correct hierarchy
    accounts = AccountModel.objects.filter(coa_model=coa).order_by('path')

    # Prepare context
    context = {
        'entity': entity,
        'accounts': accounts,
        'coa': coa,
    }

    template = get_template('djan_led/coa_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Chart_of_Accounts_{entity.slug}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF: ' + str(pisa_status.err))
    return response


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required

from decimal import Decimal
from .forms import OpeningBalanceForm


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required

from decimal import Decimal
from django.utils import timezone


@staff_member_required
def opening_balance_form(request, slug):
   
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found. Please autofill first.")
        return redirect('chart_of_accounts', slug=entity.slug)

    #accounts = AccountModel.objects.filter(coa_model=coa, depth__gt=1).order_by('code')
    #accounts = AccountModel.objects.filter(coa_model=coa).exclude(depth=1).order_by('code')
    accounts = AccountModel.objects.filter(coa_model=coa).exclude(depth=1).exclude(role='root').order_by('code')
    if request.method == 'POST':
        # Process as before
        total_debits = Decimal('0')
        total_credits = Decimal('0')
        transactions = []
        submitted = {}  # Store submitted values

        for acc in accounts:
            debit_key = f'debit_{acc.code}'
            credit_key = f'credit_{acc.code}'
            debit_val = request.POST.get(debit_key, '').strip().replace(',', '')
            credit_val = request.POST.get(credit_key, '').strip().replace(',', '')

            submitted[debit_key] = debit_val if debit_val else '0.00'
            submitted[credit_key] = credit_val if credit_val else '0.00'

            try:
                debit_amount = Decimal(debit_val) if debit_val else Decimal('0')
            except:
                debit_amount = Decimal('0')
            try:
                credit_amount = Decimal(credit_val) if credit_val else Decimal('0')
            except:
                credit_amount = Decimal('0')

            if debit_amount > 0 and credit_amount > 0:
                messages.error(request, f"Account {acc.code}: Please enter either debit OR credit, not both.")
                return render(request, 'djan_led/opening_balance_form.html', {
                    'entity': entity,
                    'accounts': accounts,
                    'submitted': submitted,
                })

            if debit_amount > 0:
                total_debits += debit_amount
                transactions.append((acc, debit_amount, 'debit'))
            elif credit_amount > 0:
                total_credits += credit_amount
                transactions.append((acc, credit_amount, 'credit'))

        if total_debits == 0 and total_credits == 0:
            messages.error(request, "Please enter at least one opening balance.")
            return render(request, 'djan_led/opening_balance_form.html', {
                'entity': entity,
                'accounts': accounts,
                'submitted': submitted,
            })

        if total_debits != total_credits:
            messages.error(request, f"Total Debits ({total_debits}) must equal Total Credits ({total_credits}).")
            return render(request, 'djan_led/opening_balance_form.html', {
                'entity': entity,
                'accounts': accounts,
                'submitted': submitted,
            })

        # Create Journal Entry...
        ledger = LedgerModel.objects.filter(entity=entity).first()
        if not ledger:
            ledger = LedgerModel.objects.create(entity=entity, name='Default Ledger')

        je = JournalEntryModel.objects.create(
            
            ledger=ledger,
            created=request.POST.get('date', '2026-01-01'),
            description='Opening balances',
            posted=False,
        )

        for acc, amount, tx_type in transactions:
            TransactionModel.objects.create(
                journal_entry=je,
                account=acc,
                amount=amount,
                tx_type=tx_type,
            )

        je.posted = True
        je.save()

        messages.success(request, f" Opening balances posted! Journal Entry #{je.uuid}")
        return redirect('djan_led:entity_dashboard', slug=entity.slug)

    # GET request
    context = {
        'entity': entity,
        'accounts': accounts,
        'today': timezone.now().date(),
        'submitted': {},  # empty for GET
    }
    return render(request, 'djan_led/opening_balance_form.html', context)


@login_required
def opening_balance_PDF(request, slug):
  
    entity = get_object_or_404(EntityModel, slug=slug)
    # security check (same as above)
    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        messages.error(request, "Access denied.")
        return redirect('entity_dashboard', slug=entity.slug)

    # Find the opening balance journal entry
    je = JournalEntryModel.objects.filter(
        ledger__entity=entity,
        description='Opening balances'
    ).order_by('-timestamp').first()

    if not je:
        messages.error(request, "No opening balance entry found.")
        return redirect('djan_led:entity_dashboard', slug=entity.slug)

    transactions = TransactionModel.objects.filter(journal_entry=je)

    template = get_template('djan_led/opening_balance_pdf.html')
    html = template.render({
        'entity': entity,
        'journal_entry': je,
        'transactions': transactions,
        'now': timezone.now(),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Opening_Balance_{entity.slug}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

from openpyxl import Workbook
from openpyxl.styles import Font
@login_required
def opening_balance_excel(request, slug):
  
    entity = get_object_or_404(EntityModel, slug=slug)
    # security check (same)
    je = JournalEntryModel.objects.filter(ledger__entity=entity, description='Opening balances').order_by('-timestamp').first()
    if not je:
        messages.error(request, "No opening balance entry found.")
        return redirect('entity_dashboard', slug=entity.slug)

    transactions = TransactionModel.objects.filter(journal_entry=je)

    wb = Workbook()
    ws = wb.active
    ws.title = "Opening Balance"
    headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    for row, tx in enumerate(transactions, start=2):
        ws.cell(row=row, column=1, value=tx.account.code)
        ws.cell(row=row, column=2, value=tx.account.name)
        if tx.tx_type == 'debit':
            ws.cell(row=row, column=3, value=float(tx.amount))
        else:
            ws.cell(row=row, column=4, value=float(tx.amount))
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Opening_Balance_{entity.slug}.xlsx"'
    wb.save(response)
    return response

@login_required
def journal_entry_detail(request, slug, pk):
   
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    
    # Access check...
    entry = get_object_or_404(JournalEntryModel, ledger__entity=entity, uuid=pk)
    transactions = TransactionModel.objects.filter(journal_entry=entry)
    total_debits = sum(tx.amount for tx in transactions if tx.tx_type == 'debit')
    total_credits = sum(tx.amount for tx in transactions if tx.tx_type == 'credit')
    context = {
        'entity': entity,
        'entry': entry,
        'transactions': transactions,
        'total_debits': total_debits,
        'total_credits': total_credits,
    }
    return render(request, 'djan_led/journal_entry_detail.html', context)

@login_required
def account_visibility(request, slug):
   
    entity = get_object_or_404(EntityModel, slug=slug)
    
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    

    try:
        profile = request.user.djan_led_profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        messages.error(request, "Access denied.")
        return redirect('entity_dashboard', slug=entity.slug)

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found. Please autofill first.")
        return redirect('chart_of_accounts', slug=entity.slug)

    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')

    if request.method == 'POST':
        action = request.POST.get('action')
        selected_uuids = [uuid for uuid in request.POST.getlist('selected_ids') if uuid.strip()]
    #    selected_uuids = request.POST.getlist('selected_ids')
        if action == 'activate':
            count = AccountModel.objects.filter(uuid__in=selected_uuids).update(active=True)
            messages.success(request, f"{count} accounts activated.")
        elif action == 'deactivate':
            count = AccountModel.objects.filter(uuid__in=selected_uuids).update(active=False)
            messages.success(request, f"{count} accounts deactivated.")
        return redirect('djan_led:account_visibility', slug=entity.slug)

    context = {
        'entity': entity,
        'accounts': accounts,
    }
    return render(request, 'djan_led/account_visibility.html', context)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

@login_required
def chart_of_accounts_excel(request, slug):
   
    entity = get_object_or_404(EntityModel, slug=slug)
    # Access check (same as other views)

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect('entity_dashboard', slug=entity.slug)

    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')

    wb = Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    # Headers
    headers = ['Code', 'Account Name', 'Role', 'Balance Type', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, acc in enumerate(accounts, start=2):
        ws.cell(row=row, column=1, value=acc.code)
        ws.cell(row=row, column=2, value=acc.name)
        ws.cell(row=row, column=3, value=acc.role.capitalize())
        ws.cell(row=row, column=4, value=acc.balance_type.capitalize())
        ws.cell(row=row, column=5, value="Active" if acc.is_active else "Inactive")

    # Auto‑width columns
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20  # A=65, B=66...

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Chart_of_Accounts_{entity.slug}.xlsx"'
    wb.save(response)
    return response


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from .models import EntityConfig
from .forms import EntityConfigForm
from django_ledger.models import EntityModel


@staff_member_required
def entity_settings(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    config, created = EntityConfig.objects.get_or_create(entity=entity)
    if request.method == "POST":
        form = EntityConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Settings updated successfully.")
            return redirect("djan_led:entity_settings", slug=entity.slug)
    else:
        form = EntityConfigForm(instance=config)
    context = {
        "entity": entity,
        "form": form,
    }
    return render(request, "djan_led/entity_config_settings.html", context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django_ledger.models import EntityModel, AccountModel
from .models import UserProfile

@login_required
def account_preferences(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    profile = request.user.djan_led_profile

    # Ensure user has access to this entity
    if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
        messages.error(request, "You don't have access to this entity.")
        return redirect('djan_led:entity_dashboard', slug=slug)

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect('djan_led:entity_dashboard', slug=slug)

    # Get all active, non‑root accounts for this entity
    all_accounts = AccountModel.objects.filter(
        coa_model=coa,
        active=True,
        depth__gt=1
    ).order_by('code')

    # Get the user's saved preference for this entity
    prefs = profile.account_preferences.get(slug, [])
    # Ensure only valid codes are kept
    valid_codes = set(acc.code for acc in all_accounts)
    prefs = [code for code in prefs if code in valid_codes]

    if request.method == 'POST':
        selected_codes = request.POST.getlist('account_codes')
        # Filter only codes that exist in this entity's COA
        selected_codes = [code for code in selected_codes if code in valid_codes]
        profile.account_preferences[slug] = selected_codes
        profile.save()
        messages.success(request, "Account preferences updated.")
        return redirect('djan_led:account_preferences', slug=slug)

    context = {
        'entity': entity,
        'all_accounts': all_accounts,
        'selected_codes': prefs,
    }
    return render(request, 'djan_led/account_preferences.html', context)


@login_required
def manual_journal_entry(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)

    if not user_can_access_entity(request.user, entity):
        return render(request, "djan_led/access_denied.html", {"entity": entity})

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("djan_led:entity_dashboard", slug=slug)

    # Get active, non-root accounts for dropdown
    accounts = AccountModel.objects.filter(
        coa_model=coa, active=True, depth__gt=1
    ).order_by("code")

    if request.method == "POST":
        date = request.POST.get("date")
        description = request.POST.get("description", "")
        debit_account_code = request.POST.get("debit_account")
        credit_account_code = request.POST.get("credit_account")
        amount = Decimal(request.POST.get("amount", "0"))

        if not all([date, debit_account_code, credit_account_code, amount]):
            messages.error(request, "All fields are required.")
            return render(
                request,
                "djan_led/manual_journal_entry.html",
                {"entity": entity, "accounts": accounts},
            )

        try:
            debit_account = AccountModel.objects.get(
                coa_model=coa, code=debit_account_code, active=True
            )
            credit_account = AccountModel.objects.get(
                coa_model=coa, code=credit_account_code, active=True
            )
        except AccountModel.DoesNotExist:
            messages.error(request, "Invalid account selected.")
            return render(
                request,
                "djan_led/manual_journal_entry.html",
                {"entity": entity, "accounts": accounts},
            )

        # Get or create ledger
        ledger = LedgerModel.objects.filter(entity=entity).first()
        if not ledger:
            ledger = LedgerModel.objects.create(entity=entity, name="Default Ledger")

        # Create Journal Entry
        je = JournalEntryModel.objects.create(
            ledger=ledger,
            timestamp=datetime.strptime(date, "%Y-%m-%d").date(),
            description=description,
            posted=False,
        )

        # Debit transaction
        TransactionModel.objects.create(
            journal_entry=je, account=debit_account, amount=amount, tx_type="debit"
        )

        # Credit transaction
        TransactionModel.objects.create(
            journal_entry=je, account=credit_account, amount=amount, tx_type="credit"
        )

        # Post the entry
        je.posted = True
        je.save()

        messages.success(request, f"Journal Entry #{je.uuid} posted successfully.")
        return redirect("djan_led:manual_journal_entry", slug=slug)

    context = {
        "entity": entity,
        "accounts": accounts,
        "today": timezone.now().date(),
    }
    return render(request, "djan_led/manual_journal_entry.html", context)


@login_required
def manual_journal_entry_create(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect('djan_led:entity_dashboard', slug=slug)

    accounts = AccountModel.objects.filter(coa_model=coa, active=True, depth__gt=1).order_by('code')

    if request.method == 'POST':
        date = request.POST.get('date')
        description = request.POST.get('description')
        debit_account_code = request.POST.get('debit_account')
        credit_account_code = request.POST.get('credit_account')
        amount = Decimal(request.POST.get('amount', '0'))

        if not all([date, description, debit_account_code, credit_account_code, amount]):
            messages.error(request, "All fields are required.")
            return render(request, 'djan_led/manual_journal_entry_create.html', {'entity': entity, 'accounts': accounts})

        # Validate accounts exist
        try:
            debit_acc = AccountModel.objects.get(coa_model=coa, code=debit_account_code, active=True)
            credit_acc = AccountModel.objects.get(coa_model=coa, code=credit_account_code, active=True)
        except AccountModel.DoesNotExist:
            messages.error(request, "Invalid account selected.")
            return render(request, 'djan_led/manual_journal_entry_create.html', {'entity': entity, 'accounts': accounts})

        # Save draft entry
        entry = ManualJournalEntry.objects.create(
            entity=entity,
            date=datetime.strptime(date, '%Y-%m-%d').date(),
            description=description,
            debit_account_code=debit_account_code,
            credit_account_code=credit_account_code,
            amount=amount,
            created_by=request.user,
            journal_status='PENDING',
        )

        messages.success(request, f"Entry #{entry.id} saved as draft. Ready for review.")
        return redirect('djan_led:manual_journal_entry_list', slug=slug)

    context = {
        'entity': entity,
        'accounts': accounts,
        'today': timezone.now().date(),
    }
    return render(request, 'djan_led/manual_journal_entry_create.html', context)


@login_required
def manual_journal_entry_list(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    entries = ManualJournalEntry.objects.filter(entity=entity).order_by("-created_at")

    if request.method == "POST":
        action = request.POST.get("action")
        selected_ids = request.POST.getlist("selected_ids")

        if action == "post":
            posted_count = 0
            for entry_id in selected_ids:
                entry = get_object_or_404(
                    ManualJournalEntry,
                    id=entry_id,
                    entity=entity,
                    journal_status="PENDING",
                )
                # Post to ledger
                result = post_manual_journal_entry(entry)
                if result:
                    posted_count += 1
                    entry.journal_status = "POSTED"
                    entry.journal_entry_id = result
                    entry.save()
            messages.success(request, f"{posted_count} entries posted.")
        return redirect("djan_led:manual_journal_entry_list", slug=slug)

    context = {
        "entity": entity,
        "entries": entries,
    }
    return render(request, "djan_led/manual_journal_entry_list.html", context)


@login_required
def pending_journal_entries(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, "djan_led/access_denied.html", {"entity": entity})

    pending = JournalEntryModel.objects.filter(
        ledger__entity=entity, posted=False
    ).order_by("-timestamp")

    if request.method == "POST":
        selected_uuids = request.POST.getlist("selected_ids")
        if selected_uuids:
            count = JournalEntryModel.objects.filter(
                uuid__in=selected_uuids, ledger__entity=entity, posted=False
            ).update(posted=True)
            messages.success(request, f"{count} journal entries posted.")
        else:
            messages.warning(request, "No entries selected.")
        return redirect("djan_led:pending_journal_entries", slug=slug)

    context = {
        "entity": entity,
        "entries": pending,
    }
    return render(request, "djan_led/pending_journal_entries.html", context)


## ======================== Chart Of Accounts ================================
from decimal import Decimal


def get_accounts_for_type(account_type, root_nodes):
    """Return list of accounts for a given entity type."""
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, "djan_led/access_denied.html", {"entity": entity})
    root_assets, root_liabilities, root_capital, root_income, root_expenses = root_nodes
    if account_type == "church":
        return [
            ("1010", "Cash", "asset", "debit", root_assets),
            ("1020", "Bank", "asset", "debit", root_assets),
            ("1030", "Accounts Receivable", "asset", "debit", root_assets),
            ("1040", "Inventory", "asset", "debit", root_assets),
            ("1050", "Prepaid Expenses", "asset", "debit", root_assets),
            ("1060", "Office Equipment", "asset", "debit", root_assets),
            ("1070", "Buildings", "asset", "debit", root_assets),
            ("2010", "Accounts Payable", "liability", "credit", root_liabilities),
            ("2020", "Accrued Expenses", "liability", "credit", root_liabilities),
            ("2030", "Bank Loans", "liability", "credit", root_liabilities),
            ("3010", "Owner's Equity", "equity", "credit", root_capital),
            ("3020", "Retained Earnings", "equity", "credit", root_capital),
            ("4010", "Revenue", "revenue", "credit", root_income),
            ("4020", "Donations", "revenue", "credit", root_income),
            ("5010", "Cost of Goods Sold", "expense", "debit", root_expenses),
            ("6010", "Salaries Expense", "expense", "debit", root_expenses),
            ("6020", "Rent Expense", "expense", "debit", root_expenses),
            ("6030", "Utilities Expense", "expense", "debit", root_expenses),
            ("6040", "Office Supplies Expense", "expense", "debit", root_expenses),
            ("6050", "Insurance Expense", "expense", "debit", root_expenses),
        ]
    elif account_type == "pos":
        return [
            ("1010", "Cash", "asset", "debit", root_assets),
            ("1020", "Bank", "asset", "debit", root_assets),
            ("1040", "Inventory", "asset", "debit", root_assets),
            ("1060", "Office Equipment", "asset", "debit", root_assets),
            ("2010", "Accounts Payable", "liability", "credit", root_liabilities),
            ("3010", "Owner's Equity", "equity", "credit", root_capital),
            ("4010", "Sales Revenue", "revenue", "credit", root_income),
            ("5010", "Cost of Goods Sold", "expense", "debit", root_expenses),
            ("6010", "Salaries Expense", "expense", "debit", root_expenses),
            ("6020", "Rent Expense", "expense", "debit", root_expenses),
            ("6030", "Utilities Expense", "expense", "debit", root_expenses),
        ]
    elif account_type == "school":
        return [
            ("1010", "Cash", "asset", "debit", root_assets),
            ("1020", "Bank", "asset", "debit", root_assets),
            ("1030", "Accounts Receivable", "asset", "debit", root_assets),
            ("1040", "Inventory", "asset", "debit", root_assets),
            ("1050", "Prepaid Expenses", "asset", "debit", root_assets),
            ("1060", "Office Equipment", "asset", "debit", root_assets),
            ("1070", "Buildings", "asset", "debit", root_assets),
            ("2010", "Accounts Payable", "liability", "credit", root_liabilities),
            ("3010", "Owner's Equity", "equity", "credit", root_capital),
            ("4010", "Tuition Revenue", "revenue", "credit", root_income),
            ("4020", "Donations", "revenue", "credit", root_income),
            ("5010", "Cost of Goods Sold", "expense", "debit", root_expenses),
            ("6010", "Salaries Expense", "expense", "debit", root_expenses),
            ("6020", "Rent Expense", "expense", "debit", root_expenses),
            ("6030", "Utilities Expense", "expense", "debit", root_expenses),
            ("6040", "Office Supplies Expense", "expense", "debit", root_expenses),
        ]
    elif account_type == "credit_union":
        return [
            ("1010", "Cash", "asset", "debit", root_assets),
            ("1020", "Bank", "asset", "debit", root_assets),
            ("1030", "Accounts Receivable", "asset", "debit", root_assets),
            ("1040", "Inventory", "asset", "debit", root_assets),
            ("1050", "Prepaid Expenses", "asset", "debit", root_assets),
            ("1060", "Office Equipment", "asset", "debit", root_assets),
            ("1070", "Buildings", "asset", "debit", root_assets),
            ("1080", "Loan Portfolio", "asset", "debit", root_assets),
            ("2010", "Accounts Payable", "liability", "credit", root_liabilities),
            ("2020", "Member Deposits", "liability", "credit", root_liabilities),
            ("3010", "Owner's Equity", "equity", "credit", root_capital),
            ("4010", "Interest Income", "revenue", "credit", root_income),
            ("4020", "Donations", "revenue", "credit", root_income),
            ("5010", "Cost of Goods Sold", "expense", "debit", root_expenses),
            ("6010", "Salaries Expense", "expense", "debit", root_expenses),
            ("6020", "Rent Expense", "expense", "debit", root_expenses),
            ("6030", "Utilities Expense", "expense", "debit", root_expenses),
            ("6040", "Office Supplies Expense", "expense", "debit", root_expenses),
        ]
    else:
        return []


@staff_member_required
def autofill_accounts(request, slug, account_type):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("chart_of_accounts", slug=slug)

    # Get root nodes
    try:
        root_assets = AccountModel.objects.get(
            coa_model=coa, name="Asset Accounts Root Node"
        )
        root_liabilities = AccountModel.objects.get(
            coa_model=coa, name="Liability Accounts Root Node"
        )
        root_capital = AccountModel.objects.get(
            coa_model=coa, name="Capital Accounts Root Node"
        )
        root_income = AccountModel.objects.get(
            coa_model=coa, name="Income Accounts Root Node"
        )
        root_expenses = AccountModel.objects.get(
            coa_model=coa, name="Expense Accounts Root Node"
        )
    except AccountModel.DoesNotExist:
        messages.error(
            request, "Root nodes missing. Please create a Chart of Accounts first."
        )
        return redirect("chart_of_accounts", slug=slug)

    root_nodes = (
        root_assets,
        root_liabilities,
        root_capital,
        root_income,
        root_expenses,
    )
    accounts_list = get_accounts_for_type(account_type, root_nodes)
    created_count = add_accounts_to_coa(coa, accounts_list)

    messages.success(
        request, f"Added {created_count} {account_type} accounts to Chart of Accounts."
    )
    return redirect("chart_of_accounts", slug=slug)


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django_ledger.models import EntityModel, AccountModel
from .utils import get_accounts_for_type, add_accounts_to_coa


@staff_member_required
def autofill_accounts(request, slug, account_type):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found. Please create one first.")
        return redirect("chart_of_accounts", slug=slug)

    # Get root nodes
    try:
        root_assets = AccountModel.objects.get(
            coa_model=coa, name="Asset Accounts Root Node"
        )
        root_liabilities = AccountModel.objects.get(
            coa_model=coa, name="Liability Accounts Root Node"
        )
        root_capital = AccountModel.objects.get(
            coa_model=coa, name="Capital Accounts Root Node"
        )
        root_income = AccountModel.objects.get(
            coa_model=coa, name="Income Accounts Root Node"
        )
        root_expenses = AccountModel.objects.get(
            coa_model=coa, name="Expense Accounts Root Node"
        )
    except AccountModel.DoesNotExist:
        messages.error(
            request, "Root nodes missing. Please create a Chart of Accounts first."
        )
        return redirect("chart_of_accounts", slug=slug)

    root_nodes = (
        root_assets,
        root_liabilities,
        root_capital,
        root_income,
        root_expenses,
    )
    accounts_list = get_accounts_for_type(account_type, root_nodes)
    created_count = add_accounts_to_coa(coa, accounts_list)

    messages.success(
        request, f"Added {created_count} {account_type} accounts to Chart of Accounts."
    )
    return redirect("chart_of_accounts", slug=slug)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django_ledger.models import EntityModel, AccountModel
from .models import UserProfile
from .utils import user_can_access_entity


@login_required
@staff_member_required
def coa_list_manage(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found for this entity.")
        return redirect("djan_led:chart_of_accounts", slug=slug)

    accounts = AccountModel.objects.filter(coa_model=coa).order_by("code")
    context = {
        "entity": entity,
        "accounts": accounts,
        "coa": coa,
    }
    return render(request, "djan_led/coa_list_manage.html", context)


@login_required
@staff_member_required
def coa_acc_edit(request, slug, account_uuid):
    entity = get_object_or_404(EntityModel, slug=slug)
   

    account = get_object_or_404(
        AccountModel, uuid=account_uuid, coa_model__entity=entity
    )

    if request.method == "POST":
        # Process form data – manually update fields
        name = request.POST.get("name")
        code = request.POST.get("code")
        role = request.POST.get("role")
        balance_type = request.POST.get("balance_type")
        is_active = request.POST.get("is_active") == "on"

        # Update account
        account.name = name
        account.code = code
        account.role = role
        account.balance_type = balance_type
        account.active = is_active
        account.save()

        messages.success(request, f"Account {account.code} updated.")
        return redirect("djan_led:coa_list_manage", slug=slug)

    context = {
        "entity": entity,
        "account": account,
    }
    return render(request, "djan_led/coa_acc_edit.html", context)


@login_required
@staff_member_required
def coa_acc_delete(request, slug, account_uuid):
    entity = get_object_or_404(EntityModel, slug=slug)
   

    account = get_object_or_404(
        AccountModel, uuid=account_uuid, coa_model__entity=entity
    )

    if request.method == "POST":
        # Check if account has transactions
        if account.transactionmodel_set.exists():
            messages.error(
                request,
                f"Cannot delete account '{account.code}' because it has transactions.",
            )
            return redirect("djan_led:coa_list_manage", slug=slug)
        account.delete()
        messages.success(request, f"Account '{account.code}' deleted.")
        return redirect("djan_led:coa_list_manage", slug=slug)

    context = {
        "entity": entity,
        "account": account,
    }
    return render(request, "djan_led/coa_acc_delete.html", context)





@staff_member_required
def add_coa_to_entity(request):
    pass



@staff_member_required
def add_coa_to_entity(request):
    if request.method != 'POST':
        return redirect('djan_led:coa_management')

    entity_slug = request.POST.get('entity_slug')
    account_type = request.POST.get('account_type')

    if not entity_slug or not account_type:
        messages.error(request, "Missing entity or account type.")
        return redirect('djan_led:coa_management')

    entity = get_object_or_404(EntityModel, slug=entity_slug)

    # Get or create the default COA
    coa = entity.default_coa
    if not coa:
        coa = entity.create_chart_of_accounts(
            assign_as_default=True,
            commit=True,
            coa_name='Default COA'
        )
        messages.info(request, f"Created new COA for {entity.name}.")

    # Get root nodes
    try:
        root_assets = AccountModel.objects.get(coa_model=coa, name='Asset Accounts Root Node')
        root_liabilities = AccountModel.objects.get(coa_model=coa, name='Liability Accounts Root Node')
        root_capital = AccountModel.objects.get(coa_model=coa, name='Capital Accounts Root Node')
        root_income = AccountModel.objects.get(coa_model=coa, name='Income Accounts Root Node')
        root_expenses = AccountModel.objects.get(coa_model=coa, name='Expense Accounts Root Node')
    except AccountModel.DoesNotExist:
        messages.error(request, f"Root nodes missing for {entity.name}.")
        return redirect('djan_led:coa_management')

    root_nodes = (root_assets, root_liabilities, root_capital, root_income, root_expenses)

    # Get the account list for the selected type
    accounts_list = get_accounts_for_type(account_type, root_nodes)

    # Add the accounts
    created_count = add_accounts_to_coa(coa, accounts_list)

    messages.success(
        request,
        f"Added {created_count} {account_type} accounts to {entity.name}."
    )
    return redirect('djan_led:coa_management')

@staff_member_required
def set_default_coa(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.chartofaccountmodel_set.first()  # Get the first COA for this entity
    if coa:
        entity.default_coa = coa
        entity.save()
        messages.success(request, f"Set {coa.name} as default COA for {entity.name}.")
    else:
        messages.error(request, "No COA found for this entity.")
    return redirect('djan_led:coa_management')


from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required


@staff_member_required
@login_required
def user_management1(request):
    # Only technical and super_admin can access
    try:
        profile = request.user.djan_led_profile
        if profile.role not in ["technical", "super_admin"]:
            messages.error(request, "Access denied. You are not authorised.")
            return redirect("after_login_redirect")
    except:
        messages.error(request, "Profile not found.")
        return redirect("after_login_redirect")

    users = User.objects.all().order_by("username")
    user_data = []

    for u in users:
        try:
            profile = u.djan_led_profile
            role = profile.role
            default_entity = (
                profile.default_entity.name if profile.default_entity else "—"
            )
            allowed = ", ".join([e.name for e in profile.allowed_entities.all()]) or "—"
        except:
            role = "—"
            default_entity = "—"
            allowed = "—"

        user_data.append(
            {
                "user": u,
                "role": role,
                "default_entity": default_entity,
                "allowed_entities": allowed,
            }
        )

    context = {
        "user_data": user_data,
    }
    return render(request, "djan_led/user_management.html", context)


@staff_member_required
def reset_user_password1(request, user_id):
    if request.method != "POST":
        return redirect("djan_led:user_management")

    user = get_object_or_404(User, id=user_id)
    # Generate a new random password
    new_password = generate_random_password()
    #user.set_password(new_password)
    #user.save()
    
    #new_password = User.objects.make_random_password(length=12)
    user.set_password(new_password)
    user.save()

    # Optionally, you could send an email with the new password.
    messages.success(
        request, f"Password for {user.username} has been reset to: {new_password}"
    )
    return redirect("djan_led:user_management")

from django.utils.crypto import get_random_string


import secrets
import string

def generate_random_password(length=12):
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(secrets.choice(alphabet) for _ in range(length))
