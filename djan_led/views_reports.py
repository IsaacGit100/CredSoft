from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import datetime

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required

#from django_ledger.io.cursor import JournalEntryCursor   # <-- Correct import
from django.db.models import Sum
from django.utils import timezone
from .models import UserProfile
from .utils import user_can_access_entity


from django_ledger.models import EntityModel, JournalEntryModel, TransactionModel, AccountModel
 
@login_required
def trial_balance_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    # Access check (same as trial_balance view)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    coa = entity.get_default_coa()
    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')

    trial_data = []
    for acc in accounts:
        debits = TransactionModel.objects.filter(account=acc, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        credits = TransactionModel.objects.filter(account=acc, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        balance = debits - credits
        if debits or credits:
            trial_data.append({
                'code': acc.code,
                'name': acc.name,
                'debit': debits,
                'credit': credits,
                'balance': balance,
            })

    total_debits = sum(item['debit'] for item in trial_data)
    total_credits = sum(item['credit'] for item in trial_data)

    context = {
        'entity': entity,
        'trial_data': trial_data,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'date': datetime.date.today(),
    }
    template = get_template('djan_led/trial_balance_pdf.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Trial_Balance_{entity.slug}_{datetime.date.today()}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response

@login_required
def trial_balance_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    coa = entity.get_default_coa()
    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Headers
    headers = ['Code', 'Account Name', 'Debit', 'Credit', 'Balance']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row = 2
    for item in trial_data:
        ws.cell(row=row, column=1, value=item['code'])
        ws.cell(row=row, column=2, value=item['name'])
        ws.cell(row=row, column=3, value=float(item['debit']))
        ws.cell(row=row, column=4, value=float(item['credit']))
        ws.cell(row=row, column=5, value=float(item['balance']))
        row += 1

    # Totals
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=float(total_debits)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_credits)).font = Font(bold=True)

    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Trial_Balance_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response


@login_required
def income_statement_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})

    # Same logic as income_statement view
    revenue_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='revenue')
    expense_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='expense')
    revenue_total = TransactionModel.objects.filter(account__in=revenue_accounts, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    expense_total = TransactionModel.objects.filter(account__in=expense_accounts, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    net_income = revenue_total - expense_total

    context = {
        'entity': entity,
        'revenue_total': revenue_total,
        'expense_total': expense_total,
        'net_income': net_income,
        'date': datetime.date.today(),
    }
    template = get_template('djan_led/income_statement_pdf.html')
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Income_Statement_{entity.slug}_{datetime.date.today()}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response


# ============================================================
# HELPER: Generate PDF from HTML
# ============================================================
def render_pdf(template_name, context, filename):
    template = get_template(template_name)
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF')
    return response

# ============================================================
# TRIAL BALANCE - PDF & EXCEL
# ============================================================
@login_required
def trial_balance_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    coa = entity.get_default_coa()
    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')
    trial_data = []
    for acc in accounts:
        debits = TransactionModel.objects.filter(account=acc, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        credits = TransactionModel.objects.filter(account=acc, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        if debits or credits:
            trial_data.append({'code': acc.code, 'name': acc.name, 'debit': debits, 'credit': credits})
    total_debits = sum(i['debit'] for i in trial_data)
    total_credits = sum(i['credit'] for i in trial_data)
    context = {'entity': entity, 'trial_data': trial_data, 'total_debits': total_debits, 'total_credits': total_credits, 'date': datetime.date.today()}
    return render_pdf('djan_led/trial_balance_pdf.html', context, f"Trial_Balance_{entity.slug}_{datetime.date.today()}")

@login_required
def trial_balance_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    coa = entity.get_default_coa()
    accounts = AccountModel.objects.filter(coa_model=coa).exclude(role='root').order_by('code')
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    headers = ['Code', 'Account Name', 'Debit', 'Credit']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
    row = 2
    for acc in accounts:
        debits = TransactionModel.objects.filter(account=acc, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        credits = TransactionModel.objects.filter(account=acc, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        if debits or credits:
            ws.cell(row=row, column=1, value=acc.code)
            ws.cell(row=row, column=2, value=acc.name)
            ws.cell(row=row, column=3, value=float(debits))
            ws.cell(row=row, column=4, value=float(credits))
            row += 1
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Trial_Balance_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# INCOME STATEMENT - PDF & EXCEL
# ============================================================
@login_required
def income_statement_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    revenue_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='revenue')
    expense_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='expense')
    revenue_total = TransactionModel.objects.filter(account__in=revenue_accounts, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    expense_total = TransactionModel.objects.filter(account__in=expense_accounts, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    net_income = revenue_total - expense_total
    context = {'entity': entity, 'revenue_total': revenue_total, 'expense_total': expense_total, 'net_income': net_income, 'date': datetime.date.today()}
    return render_pdf('djan_led/income_statement_pdf.html', context, f"Income_Statement_{entity.slug}_{datetime.date.today()}")

@login_required
def income_statement_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    revenue_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='revenue')
    expense_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='expense')
    revenue_total = TransactionModel.objects.filter(account__in=revenue_accounts, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    expense_total = TransactionModel.objects.filter(account__in=expense_accounts, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    net_income = revenue_total - expense_total
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"
    ws['A1'] = "Income Statement"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = entity.name
    ws['A3'] = f"Generated: {datetime.date.today()}"
    ws['A5'] = "Revenue"
    ws['A5'].font = Font(bold=True)
    ws['A6'] = revenue_total
    ws['A8'] = "Expenses"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = expense_total
    ws['A11'] = "Net Income"
    ws['A11'].font = Font(bold=True)
    ws['B11'] = net_income
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 25
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Income_Statement_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# BALANCE SHEET - PDF & EXCEL
# ============================================================
def get_bal(entity):
    assets = AccountModel.objects.filter(coa_model__entity=entity, role='asset')
    liabilities = AccountModel.objects.filter(coa_model__entity=entity, role='liability')
    equity = AccountModel.objects.filter(coa_model__entity=entity, role='equity')
    def bal(acc):
        d = TransactionModel.objects.filter(account=acc, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        c = TransactionModel.objects.filter(account=acc, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        return d - c if acc.balance_type == 'debit' else c - d
    return {'assets': sum(bal(a) for a in assets), 'liabilities': sum(bal(l) for l in liabilities), 'equity': sum(bal(e) for e in equity)}

@login_required
def balance_sheet_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    b = get_bal(entity)
    context = {'entity': entity, 'assets': b['assets'], 'liabilities': b['liabilities'], 'equity': b['equity'], 'date': datetime.date.today()}
    return render_pdf('djan_led/balance_sheet_pdf.html', context, f"Balance_Sheet_{entity.slug}_{datetime.date.today()}")

@login_required
def balance_sheet_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    b = get_bal(entity)
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws['A1'] = "Balance Sheet"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = entity.name
    ws['A3'] = f"Generated: {datetime.date.today()}"
    ws['A5'] = "Assets"
    ws['A5'].font = Font(bold=True)
    ws['A6'] = b['assets']
    ws['A8'] = "Liabilities"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = b['liabilities']
    ws['A11'] = "Equity"
    ws['A11'].font = Font(bold=True)
    ws['A12'] = b['equity']
    ws['A14'] = "Total Liabilities + Equity"
    ws['A14'].font = Font(bold=True)
    ws['B14'] = b['liabilities'] + b['equity']
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Balance_Sheet_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# JOURNAL ENTRIES - PDF & EXCEL
# ============================================================
@login_required
def journal_entries_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    entries = JournalEntryModel.objects.filter(ledger__entity=entity).order_by('-timestamp')[:100]
    data = []
    for je in entries:
        txs = TransactionModel.objects.filter(journal_entry=je)
        data.append({'je': je, 'txs': txs})
    context = {'entity': entity, 'data': data, 'date': datetime.date.today()}
    return render_pdf('djan_led/journal_entries_pdf.html', context, f"Journal_Entries_{entity.slug}_{datetime.date.today()}")

@login_required
def journal_entries_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    entries = JournalEntryModel.objects.filter(ledger__entity=entity).order_by('-timestamp')[:100]
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"
    headers = ['Date', 'Description', 'Account', 'Debit', 'Credit', 'Status']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')
    row = 2
    for je in entries:
        txs = TransactionModel.objects.filter(journal_entry=je)
        for tx in txs:
            ws.cell(row=row, column=1, value=je.timestamp.strftime('%Y-%m-%d') if je.timestamp else '')
            ws.cell(row=row, column=2, value=je.description or '')
            ws.cell(row=row, column=3, value=f"{tx.account.code} - {tx.account.name}")
            ws.cell(row=row, column=4, value=float(tx.amount) if tx.tx_type == 'debit' else 0)
            ws.cell(row=row, column=5, value=float(tx.amount) if tx.tx_type == 'credit' else 0)
            ws.cell(row=row, column=6, value='Posted' if je.posted else 'Draft')
            row += 1
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Journal_Entries_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# CASH FLOW - PDF & EXCEL
# ============================================================
@login_required
def cash_flow_pdf(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    cash_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='asset', code__in=['1010', '1020', '1211', '1212'])
    inflows = TransactionModel.objects.filter(account__in=cash_accounts, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    outflows = TransactionModel.objects.filter(account__in=cash_accounts, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    context = {'entity': entity, 'inflows': inflows, 'outflows': outflows, 'net': inflows - outflows, 'date': datetime.date.today()}
    return render_pdf('djan_led/cash_flow_pdf.html', context, f"Cash_Flow_{entity.slug}_{datetime.date.today()}")

@login_required
def cash_flow_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    if not user_can_access_entity(request.user, entity):
        return render(request, 'djan_led/access_denied.html', {'entity': entity})
    cash_accounts = AccountModel.objects.filter(coa_model__entity=entity, role='asset', code__in=['1010', '1020', '1211', '1212'])
    inflows = TransactionModel.objects.filter(account__in=cash_accounts, tx_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    outflows = TransactionModel.objects.filter(account__in=cash_accounts, tx_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    ws['A1'] = "Cash Flow Statement"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = entity.name
    ws['A3'] = f"Generated: {datetime.date.today()}"
    ws['A5'] = "Cash Inflows"
    ws['A5'].font = Font(bold=True)
    ws['A6'] = inflows
    ws['A8'] = "Cash Outflows"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = outflows
    ws['A11'] = "Net Cash Flow"
    ws['A11'].font = Font(bold=True)
    ws['B11'] = inflows - outflows
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cash_Flow_{entity.slug}_{datetime.date.today()}.xlsx"'
    wb.save(response)
    return response
