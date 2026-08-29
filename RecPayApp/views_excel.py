from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q, Count, F
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required

from django.shortcuts import render
from django.db import models

# Create your views here.
from decimal import Decimal

import re
import io

from django.utils.dateparse import parse_date
from django.contrib import messages
from django.utils import timezone
from datetime import datetime

from django.core.paginator import Paginator

from django import template
register = template.Library()

import json

## Import Tables
from .models import Trans
from MembersApp.models import Master
from UserAuth.models import User
from coa.models import ChartOfAccounts
from LoanApp.models import Loan
from django_ledger.models import EntityModel, LedgerModel, JournalEntryModel, AccountModel, TransactionModel

## Import Views
from . import views
from . import views_pdf
from . import views_excel


# from .models import Journal, Statement, Ledger, Ledger_Statement, Loan


from django.template.loader import render_to_string
import pdfkit

# Create your views here.
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.units import cm

from django import template

from django.http import HttpResponse
from django.shortcuts import render
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
import io


# views.py
from django.shortcuts import render
from django.core.paginator import Paginator

from django.contrib.auth.decorators import login_required
from reportlab.lib import colors


from datetime import datetime


from .models import Trans
from MembersApp.models import Master
from UserAuth.models import User
from coa.models import ChartOfAccounts
from LoanApp.models import Loan


def trans_all_excel(request, slug):
    """Generate Excel for ALL transactions - with clean payment details"""
    entity=get_object_or_404(EntityModel, slug=slug)
    # Get all transactions
    transactions = Trans.objects.all().order_by('-date', '-id')
    
    # Calculate totals
    total_receipts = transactions.filter(trans_type='Receipts').aggregate(total=Sum('amount'))['total'] or 0
    total_payments = transactions.filter(trans_type='Payments').aggregate(total=Sum('amount'))['total'] or 0
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = "ST. ANDREWS CO-OPERATIVE CREDIT UNION"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = "FINANCIAL TRANSACTIONS REPORT"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws[f'A{current_row}'].alignment = center_align
    current_row += 2
    
    # Summary
    ws[f'A{current_row}'] = "SUMMARY"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = "Total Records:"
    ws[f'B{current_row}'] = transactions.count()
    ws[f'C{current_row}'] = "Receipts:"
    ws[f'D{current_row}'] = float(total_receipts)
    ws[f'D{current_row}'].number_format = '"₵"#,##0.00'
    ws[f'E{current_row}'] = "Payments:"
    ws[f'F{current_row}'] = float(total_payments)
    ws[f'F{current_row}'].number_format = '"₵"#,##0.00'
    ws[f'G{current_row}'] = "Net:"
    ws[f'H{current_row}'] = float(total_receipts - total_payments)
    ws[f'H{current_row}'].number_format = '"₵"#,##0.00'
    current_row += 3
    
    # Headers
    headers = ['Date', 'ID', 'Reference', 'Mem ID', 'Member/Name', 
               'Code', 'Ledger', 'Details', 'Receipts', 'Payments', 'Payment Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
    current_row += 1
    
    # Data rows
    for t in transactions:
        # Get member/name
        if t.member:
            member_id = t.member.id
            person = t.member_name or f"{t.member.first_name} {t.member.last_name}"
        elif t.non_member_name:
            member_id = '-'
            person = t.non_member_name
        else:
            member_id = '-'
            person = '-'
        
        # Get CLEAN payment details
        payment_details = ''
        if t.pay_mode == 'Cheque':
            details = []
            if t.cheque_no:
                details.append(f"{t.cheque_no}")
            if t.cheque_date:
                details.append(f"{t.cheque_date.strftime('%d/%m/%Y')}")
            if t.bank:
                details.append(f"{t.bank}")
            if t.bank_branch:
                details.append(f"{t.bank_branch}")
            if t.bank_no:
                details.append(f"{t.bank_no}")
            payment_details = ' : '.join(details) if details else 'CASH'
        elif t.pay_mode == 'Transfer':
            details = []
            if t.momo_no:
                details.append(f"{t.momo_no}")
            if t.momo_name:
                details.append(f"{t.momo_name}")
            payment_details = ' : '.join(details) if details else 'CASH'
        else:
            payment_details = 'CASH'
        
        # Write data
        ws.cell(row=current_row, column=1, value=t.date.strftime('%d/%m/%Y')).alignment = center_align
        ws.cell(row=current_row, column=2, value=t.id).alignment = center_align
        ws.cell(row=current_row, column=3, value=t.rec_vou_no or '-').alignment = center_align
        ws.cell(row=current_row, column=4, value=member_id).alignment = center_align
        ws.cell(row=current_row, column=5, value=person).alignment = left_align
        ws.cell(row=current_row, column=6, value=t.ledger_code or '-').alignment = left_align
        ws.cell(row=current_row, column=7, value=t.ledger_name or '-').alignment = left_align
        ws.cell(row=current_row, column=8, value=t.details or '-').alignment = left_align
        
        # Amount columns
        if t.trans_type == 'Receipts':
            receipt_cell = ws.cell(row=current_row, column=9, value=float(t.amount))
            receipt_cell.number_format = '"₵"#,##0.00'
            receipt_cell.alignment = right_align
        else:
            payment_cell = ws.cell(row=current_row, column=10, value=float(t.amount))
            payment_cell.number_format = '"₵"#,##0.00'
            payment_cell.alignment = right_align
        
        ws.cell(row=current_row, column=11, value=payment_details).alignment = left_align
        
        current_row += 1
    
    # Totals row
    ws.cell(row=current_row, column=1, value='TOTAL').font = bold_font
    
    total_receipts_cell = ws.cell(row=current_row, column=9, value=float(total_receipts))
    total_receipts_cell.font = bold_font
    total_receipts_cell.number_format = '"₵"#,##0.00'
    total_receipts_cell.alignment = right_align
    
    total_payments_cell = ws.cell(row=current_row, column=10, value=float(total_payments))
    total_payments_cell.font = bold_font
    total_payments_cell.number_format = '"₵"#,##0.00'
    total_payments_cell.alignment = right_align
    
    net_cell = ws.cell(row=current_row, column=11, value=f"NET: ₵{total_receipts - total_payments:,.2f}")
    net_cell.font = bold_font
    
    # Adjust column widths
    column_widths = [10, 6, 12, 8, 25, 8, 20, 20, 12, 12, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="transactions_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

def trans_excel1(request, pk):  # Make sure it accepts 'pk'
    """Generate Excel for a single transaction"""
    
    # Get the transaction
    trans = get_object_or_404(Trans, pk=pk)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Transaction_{trans.id}"
    
    # Styles
    bold_font = Font(bold=True)
    normal_font = Font()
    header_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "ST. ANDREWS CO-OPERATIVE CREDIT UNION"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "TRANSACTION RECEIPT"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    ws[f'A{current_row}'].alignment = center_align
    current_row += 2
    
    # Reference
    ws[f'A{current_row}'] = f"Reference: {trans.rec_vou_no or trans.trans_no}"
    ws[f'C{current_row}'] = f"Date: {trans.date.strftime('%d/%m/%Y')}"
    current_row += 2
    
    # Transaction Details
    ws[f'A{current_row}'] = "TRANSACTION DETAILS"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = "Transaction Type:"
    ws[f'B{current_row}'] = trans.trans_type
    ws[f'C{current_row}'] = "Payment Mode:"
    ws[f'D{current_row}'] = trans.pay_mode
    current_row += 1
    
    ws[f'A{current_row}'] = "Amount:"
    ws[f'B{current_row}'] = float(trans.amount)
    ws[f'B{current_row}'].number_format = '"₵"#,##0.00'
    current_row += 2
    
    # Person Information
    ws[f'A{current_row}'] = "PERSON INFORMATION"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    if trans.member:
        ws[f'A{current_row}'] = "Member ID:"
        ws[f'B{current_row}'] = trans.member.id
        current_row += 1
        ws[f'A{current_row}'] = "Member Name:"
        ws[f'B{current_row}'] = trans.member_name
    elif trans.non_member_name:
        ws[f'A{current_row}'] = "Non-Member Name:"
        ws[f'B{current_row}'] = trans.non_member_name
        current_row += 1
        if trans.non_member_contact:
            ws[f'A{current_row}'] = "Contact:"
            ws[f'B{current_row}'] = trans.non_member_contact
    else:
        ws[f'A{current_row}'] = "No person information available"
    current_row += 2
    
    # Ledger Information
    ws[f'A{current_row}'] = "LEDGER INFORMATION"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = "Ledger Code:"
    ws[f'B{current_row}'] = trans.ledger_code or '-'
    current_row += 1
    ws[f'A{current_row}'] = "Ledger Name:"
    ws[f'B{current_row}'] = trans.ledger_name or '-'
    current_row += 1
    if trans.details:
        ws[f'A{current_row}'] = "Details:"
        ws[f'B{current_row}'] = trans.details
        current_row += 1
    current_row += 1
    
    # Payment Details
    if trans.pay_mode == 'Cheque' and (trans.cheque_no or trans.bank):
        ws[f'A{current_row}'] = "CHEQUE DETAILS"
        ws[f'A{current_row}'].font = bold_font
        current_row += 1
        
        if trans.cheque_no:
            ws[f'A{current_row}'] = "Cheque No:"
            ws[f'B{current_row}'] = trans.cheque_no
            current_row += 1
        if trans.cheque_date:
            ws[f'A{current_row}'] = "Cheque Date:"
            ws[f'B{current_row}'] = trans.cheque_date.strftime('%d/%m/%Y')
            current_row += 1
        if trans.bank:
            ws[f'A{current_row}'] = "Bank:"
            ws[f'B{current_row}'] = trans.bank
            current_row += 1
        if trans.bank_branch:
            ws[f'A{current_row}'] = "Branch:"
            ws[f'B{current_row}'] = trans.bank_branch
            current_row += 1
        if trans.bank_no:
            ws[f'A{current_row}'] = "Account No:"
            ws[f'B{current_row}'] = trans.bank_no
            current_row += 1
            
    elif trans.pay_mode == 'Transfer' and (trans.momo_no or trans.momo_name):
        ws[f'A{current_row}'] = "TRANSFER DETAILS"
        ws[f'A{current_row}'].font = bold_font
        current_row += 1
        
        if trans.momo_no:
            ws[f'A{current_row}'] = "Mobile Money No:"
            ws[f'B{current_row}'] = trans.momo_no
            current_row += 1
        if trans.momo_name:
            ws[f'A{current_row}'] = "Account Name:"
            ws[f'B{current_row}'] = trans.momo_name
            current_row += 1
    
    current_row += 2
    
    # Footer
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "-" * 60
    current_row += 2
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y at %H:%M')}"
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f"Transaction ID: {trans.id}"
    ws[f'A{current_row}'].alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"transaction_{trans.rec_vou_no or trans.id}_{trans.date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# RecPayApp/views.py
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from .models import Trans

@login_required
def trans_excel(request, slug, pk):
    """Export single transaction as Excel"""
    entity=get_object_or_404(EntityModel, slug=slug)
    transaction = get_object_or_404(Trans, pk=pk)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Transaction_{transaction.trans_no}"
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "ST. ANDREWS CO-OPERATIVE CREDIT UNION"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:B2')
    ws['A2'] = "Transaction Receipt"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Transaction Details
    row = 4
    details = [
        ("Transaction Number:", transaction.trans_no),
        ("Receipt/Voucher No:", transaction.rec_vou_no),
        ("Date:", transaction.date.strftime('%d/%m/%Y')),
        ("Type:", transaction.trans_type),
        ("Amount:", f"₵{transaction.amount:,.2f}"),
        ("Payment Mode:", transaction.pay_mode),
        ("Status:", transaction.status),
    ]
    
    for label, value in details:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Party Information
    ws[f'A{row}'] = "Party Information"
    ws[f'A{row}'].font = header_font
    row += 1
    
    party_data = [
        ("Name Type:", "Member" if transaction.member else "Non-Member"),
        ("Name:", transaction.member.full_name if transaction.member else transaction.non_member_name),
    ]
    
    if transaction.member:
        party_data.append(("Member ID:", transaction.member.id))
        party_data.append(("Phone:", transaction.member.telephone1 or "-"))
    
    for label, value in party_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    row += 1
    
    # Accounting Information
    ws[f'A{row}'] = "Accounting Information"
    ws[f'A{row}'].font = header_font
    row += 1
    
    accounting_data = [
        ("Ledger Account:", f"{transaction.ledger_code} - {transaction.ledger_name}"),
        ("Description:", transaction.details or "-"),
    ]
    
    for label, value in accounting_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Transaction_{transaction.trans_no}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
