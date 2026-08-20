from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from . models import  OpeningBalanceLine
from FinanceApp.models import GeneralLedger
from coa.models import ChartOfAccounts
from django.utils import timezone
from django.contrib import messages
from decimal import Decimal
from . forms import OpeningBalanceLineForm


# Create your views here.


def opening_balance_home(request, slug):
    return render(request, 'OpenBals/opening_balance_home.html')

@login_required
def opening_balance_list(request, slug):
    lines = OpeningBalanceLine.objects.select_related('account', 'created_by').all()
    context = {
        'lines': lines,
        'pending_count': OpeningBalanceLine.objects.filter(status='PENDING').count(),
        'approved_count': OpeningBalanceLine.objects.filter(status='APPROVED').count(),
        'posted_count': OpeningBalanceLine.objects.filter(status='POSTED').count(),
        'today': timezone.now().date(),
        'session_default_date': request.session.get('default_opening_balance_date'),
    }
    return render(request, 'OpenBals/opening_balance_list.html', context)



@login_required
def opening_balance_list1(request, slug):
    lines = OpeningBalanceLine.objects.select_related('account', 'created_by').all()
    context = {
        'lines': lines,
        'pending_count': OpeningBalanceLine.objects.filter(status='PENDING').count(),
        'approved_count': OpeningBalanceLine.objects.filter(status='APPROVED').count(),
        'posted_count': OpeningBalanceLine.objects.filter(status='POSTED').count(),
    }
    return render(request, 'OpenBals/opening_balance_list.html', context)


# ---------- ADD ----------
@login_required
def opening_balance_create(request, slug):
    # Get the default date from session, or use today
    default_date = request.session.get('default_opening_balance_date')
    if default_date:
        # Convert string to date object (or keep as string; form field accepts string)
        from datetime import datetime
        try:
            default_date = datetime.strptime(default_date, '%Y-%m-%d').date()
        except ValueError:
            default_date = timezone.now().date()
    else:
        default_date = timezone.now().date()

    if request.method == 'POST':
        form = OpeningBalanceLineForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.created_by = request.user
            instance.save()
            messages.success(request, "Opening balance added.")
            return redirect('OpenBals:list')
    else:
        # Pre‑fill date with the session value or today
        initial = {'date': default_date}
        form = OpeningBalanceLineForm(initial=initial)
    
    return render(request, 'opening_balances/form.html', {'form': form})

@login_required
def clear_default_date(request, slug):
    if 'default_opening_balance_date' in request.session:
        del request.session['default_opening_balance_date']
    messages.info(request, "Default date cleared. New entries will use today's date.")
    return redirect('OpenBals:list')


@login_required
def opening_balance_create1(request, slug):
    if request.method == 'POST':
        form = OpeningBalanceLineForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.created_by = request.user
            instance.save()
            messages.success(request, "Opening balance added.")
            return redirect('OpenBals:opening_balance_list')
    else:
        form = OpeningBalanceLineForm()
    return render(request, 'OpenBals/opening_balance_form.html', {'form': form})


# ---------- EDIT ----------
@login_required
def opening_balance_edit(request, slug, pk):
    line = get_object_or_404(OpeningBalanceLine, pk=pk)
    # Only allow editing if PENDING (or you can allow APPROVED too, up to you)
    if line.status != 'PENDING':
        messages.error(request, "Only pending lines can be edited.")
        return redirect('OpenBals:opening_balance_list')
    if request.method == 'POST':
        form = OpeningBalanceLineForm(request.POST, instance=line)
        if form.is_valid():
            form.save()
            messages.success(request, "Opening balance updated.")
            return redirect('OpenBals:opening_balance_list')
    else:
        form = OpeningBalanceLineForm(instance=line)
    return render(request, 'OpenBals/opening_balances_form.html', {'form': form})

# ---------- DELETE ----------
@login_required
def opening_balance_delete(request, slug, pk):
    line = get_object_or_404(OpeningBalanceLine, pk=pk)
    if line.status != 'PENDING':
        messages.error(request, "Only pending lines can be deleted.")
        return redirect('OpenBals:opening_balance_list')
    if request.method == 'POST':
        line.delete()
        messages.success(request, "Opening balance deleted.")
        return redirect('OpenBals:opening_balance_list')
    return render(request, 'OpenBals/opening_balance_confirm_delete.html', {'object': line})

# ---------- BULK APPROVE ----------
@login_required
def opening_balance_approve(request, slug):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        if not ids:
            messages.error(request, "No items selected.")
            return redirect('opening_balances:list')
        lines = OpeningBalanceLine.objects.filter(id__in=ids, status='PENDING')
        count = lines.update(
            status='APPROVED',
            approved_by=request.user,
            approved_at=timezone.now()
        )
        messages.success(request, f"{count} line(s) approved.")
    return redirect('OpenBals:opening_balance_list')

# ---------- POSTING PAGE (separate view) ----------
@login_required
def opening_balance_post_page(request, slug):
    # Only show approved lines that are not yet posted
    lines = OpeningBalanceLine.objects.filter(status='APPROVED').select_related('account')
    return render(request, 'OpenBals/opening_balance_post.html', {'lines': lines})

# ---------- BULK POST (to GeneralLedger) ----------
@login_required
def opening_balance_post_execute(request, slug):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        if not ids:
            messages.error(request, "No items selected.")
            return redirect('OpenBals:post_page')
        lines = OpeningBalanceLine.objects.filter(id__in=ids, status='APPROVED')
        for line in lines:
            # Update GeneralLedger
            ledger, created = GeneralLedger.objects.get_or_create(account=line.account)
            # Ensure decimals
            if isinstance(ledger.opening_balance, float):
                ledger.opening_balance = Decimal(str(ledger.opening_balance))
            if isinstance(ledger.current_balance, float):
                ledger.current_balance = Decimal(str(ledger.current_balance))

            if line.debit:
                ledger.opening_balance += line.debit
                ledger.current_balance += line.debit
            else:
                ledger.opening_balance -= line.credit
                ledger.current_balance -= line.credit

            ledger.open_bal_date = line.date
            ledger.save()

            # Mark line as posted
            line.status = 'POSTED'
            line.posted_by = request.user
            line.posted_at = timezone.now()
            line.save()

        messages.success(request, f"{len(lines)} line(s) posted to General Ledger.")
    return redirect('OpenBals:opening_balance_post_page')


from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from .models import OpeningBalanceLine

@login_required
def export_pdf(request, slug):
    lines = OpeningBalanceLine.objects.select_related('account').all()
    template = get_template('OpenBals/open_bal_pdf_export.html')
    html = template.render({'lines': lines, 'user': request.user})  # ✅ pass user
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=opening_balances.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import OpeningBalanceLine

@login_required
def export_excel(request, slug):
    wb = Workbook()
    ws = wb.active
    ws.title = "Opening Balances"

    # Headers
    headers = ['Account', 'Debit', 'Credit', 'Date', 'Status', 'Created By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    lines = OpeningBalanceLine.objects.select_related('account', 'created_by').all()

    for row_idx, line in enumerate(lines, 2):
        ws.cell(row=row_idx, column=1, value=str(line.account))
        ws.cell(row=row_idx, column=2, value=float(line.debit) if line.debit else 0)
        ws.cell(row=row_idx, column=3, value=float(line.credit) if line.credit else 0)
        #  Safe date handling
        ws.cell(row=row_idx, column=4, value=line.date.strftime("%Y-%m-%d") if line.date else "")
        ws.cell(row=row_idx, column=5, value=line.get_status_display())
        #  Safe created_by handling
        if line.created_by:
            ws.cell(row=row_idx, column=6, value=line.created_by.get_full_name() or line.created_by.username)
        else:
            ws.cell(row=row_idx, column=6, value="System")

    # Auto-width
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=opening_balances.xlsx'
    wb.save(response)
    return response

from django.contrib import messages
from django.shortcuts import redirect
from django.utils import timezone

from django.shortcuts import redirect, render
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import OpeningBalanceLine

@login_required
def bulk_update_date(request, slug):
    if request.method == 'POST':
        new_date = request.POST.get('new_date')
        if not new_date:
            messages.error(request, "Please select a date.")
            return redirect('OpenBals:list')
        
        # 1. Update all existing records with this date
        count = OpeningBalanceLine.objects.update(date=new_date)
        
        # 2. Save this date in the session for future entries
        request.session['default_opening_balance_date'] = str(new_date)
        
        messages.success(request, f"Date set to {new_date} for {count} records. New records will use this date by default.")
    return redirect('OpenBals:list')











@login_required
def opening_balance_post1(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
        if not ids:
            messages.error(request, "No items selected.")
            return redirect('OpenBals:opening_balance_list')

        lines = OpeningBalanceLine.objects.filter(id__in=ids, status='APPROVED')
        for line in lines:
            # Update or create GeneralLedger
            ledger, created = GeneralLedger.objects.get_or_create(account=line.account)
            # Ensure Decimal types
            if isinstance(ledger.opening_balance, float):
                ledger.opening_balance = Decimal(str(ledger.opening_balance))
            if isinstance(ledger.current_balance, float):
                ledger.current_balance = Decimal(str(ledger.current_balance))

            if line.debit:
                ledger.opening_balance += line.debit
                ledger.current_balance += line.debit
            else:
                # credit – usually subtract from opening balance, adjust as needed
                ledger.opening_balance -= line.credit
                ledger.current_balance -= line.credit

            ledger.open_bal_date = line.date
            ledger.save()

            # Mark line as posted
            line.status = 'POSTED'
            line.posted_by = request.user
            line.posted_at = timezone.now()
            line.save()

        messages.success(request, f"{len(lines)} lines posted.")
    return redirect('OpenBals:opening_balance_list')


def opening_balance_list2(request):
    lines = OpeningBalanceLine.objects.all().order_by('-date')
    pending_count = OpeningBalanceLine.objects.filter(status='PENDING').count()
    approved_count = OpeningBalanceLine.objects.filter(status='APPROVED').count()
    posted_count = OpeningBalanceLine.objects.filter(status='POSTED').count()
    return render(request, 'OpenBals/opening_balance_list.html', {
        'lines': lines,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'posted_count': posted_count,
    })

# ---------- Batch List ----------
@login_required
def opening_balance_list2(request):
    batches = OpeningBalanceBatch.objects.select_related('created_by').all().order_by('-created_at')
    return render(request, 'OpenBals/opening_balance_list.html', {'batches': batches})

@login_required
def opening_balance_detail1(request, pk):
    batch = get_object_or_404(OpeningBalanceBatch, pk=pk)
    return render(request, 'OpenBals/batch_detail.html', {'batch': batch})

