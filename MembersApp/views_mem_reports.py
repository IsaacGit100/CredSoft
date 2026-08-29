# MembersApp/views_reports_pdf.py
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from .models import Master
from django_ledger.models import EntityModel

# ============================================================
# 1. MEMBERS INFORMATION REPORT
# ============================================================
@login_required
def members_info_pdf(request, slug):
    """PDF Report - All members basic information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=1*cm, bottomMargin=1*cm,
                           leftMargin=1*cm, rightMargin=1*cm)
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, textColor=colors.black)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=11, textColor=colors.black, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, textColor=colors.black)
    
    elements = []
    
    # Title
    elements.append(Paragraph("MEMBERS INFORMATION REPORT", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Table Data
    members = Master.objects.filter(is_deleted=False).order_by('last_name', 'first_name')
    
    table_data = [['ID', 'Full Name', 'Title', 'Gender', 'Date of Birth', 'Date Enrolled', 'Status', 'Role']]
    
    for member in members:
        table_data.append([
            str(member.id),
            member.full_name or '-',
            member.get_title_display(),
            member.get_gender_display(),
            member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '-',
            member.date_enrolled.strftime('%Y-%m-%d') if member.date_enrolled else '-',
            member.get_mem_status_display(),
            member.get_role_display(),
        ])
    
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Members_Information_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============================================================
# 2. MEMBERS CONTACT REPORT
# ============================================================
@login_required
def members_contact_pdf(request, slug):
    """PDF Report - Members contact information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=1*cm, bottomMargin=1*cm,
                           leftMargin=1*cm, rightMargin=1*cm)

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, textColor=colors.black)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, textColor=colors.black)

    elements = []

    elements.append(Paragraph("MEMBERS CONTACT REPORT", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.5*cm))

    members = Master.objects.filter(is_deleted=False).order_by('last_name', 'first_name')

    table_data = [['ID', 'Full Name', 'Phone 1', 'Phone 2', 'Email', 'City', 'Postal Address', 'Residential Address']]

    for member in members:
        table_data.append([
            str(member.id),
            member.full_name or '-',
            member.telephone1 or '-',
            member.telephone2 or '-',
            member.email_address or '-',
            member.city or '-',
            (member.postal_address or '-')[:50],
            (member.residential_address or '-')[:50],
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Members_Contact_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============================================================
# 3. NEXT OF KIN REPORT
# ============================================================
@login_required
def next_of_kin_pdf(request, slug):
    """PDF Report - Next of Kin information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=1*cm, bottomMargin=1*cm,
                           leftMargin=1*cm, rightMargin=1*cm)

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, textColor=colors.black)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, textColor=colors.black)

    elements = []

    elements.append(Paragraph("NEXT OF KIN REPORT", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.5*cm))

    members = Master.objects.filter(is_deleted=False).order_by('last_name', 'first_name')

    table_data = [['ID', 'Member Name', 'NOK Name', 'NOK Address', 'NOK Phone', 'Relation', 'Percentage']]

    for member in members:
        # First NOK
        if member.nok_name1:
            table_data.append([
                str(member.id),
                member.full_name or '-',
                member.nok_name1,
                (member.nok_address1 or '-')[:40],
                member.nok_telephone1 or '-',
                member.nok_relation1 or '-',
                f"{member.nok_percent1 or 0}%",
            ])
        # Second NOK
        if member.nok_name2:
            table_data.append([
                str(member.id),
                member.full_name or '-',
                member.nok_name2,
                (member.nok_address2 or '-')[:40],
                member.nok_telephone2 or '-',
                member.nok_relation2 or '-',
                f"{member.nok_percent2 or 0}%",
            ])
        # Third NOK
        if member.nok_name3:
            table_data.append([
                str(member.id),
                member.full_name or '-',
                member.nok_name3,
                (member.nok_address3 or '-')[:40],
                member.nok_telephone3 or '-',
                member.nok_relation3 or '-',
                f"{member.nok_percent3 or 0}%",
            ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Next_Of_Kin_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


# ============================================================
# 4. FINANCIAL REPORT
# ============================================================
@login_required
def financial_report_pdf(request, slug):
    """PDF Report - Members financial information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           topMargin=1*cm, bottomMargin=1*cm,
                           leftMargin=1*cm, rightMargin=1*cm)

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=14, alignment=TA_CENTER, textColor=colors.black)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=9, textColor=colors.black)

    elements = []

    elements.append(Paragraph("MEMBERS FINANCIAL REPORT", title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.5*cm))

    members = Master.objects.filter(is_deleted=False).order_by('last_name', 'first_name')

    table_data = [['ID', 'Full Name', 'Shares', 'Shares Withd', 'Deposits', 'Dep. Withd', 'Dividend', 'Div. Withd',
                   'Loan Balance', 'Guaranteed', 'Guaranted', 'Avail bal', 'Int_Accrued', 'Int_Deferred']]

    for member in members:
        table_data.append([
            str(member.id),
            member.full_name or '-',
            f"{(member.tot_shares if member.tot_shares else 0.00):>15,.2f}",
            f"{(member.tot_shares_withdrawal if member.tot_shares_withdrawal else 0.00):>15,.2f}",
            f"{(member.tot_deposits if member.tot_deposits else 0.00):>15,.2f}",
            f"{(member.tot_deposit_withdrawal if member.tot_deposit_withdrawal else 0.00):>15,.2f}",
            f"{(member.tot_dividend if member.tot_dividend else 0.00):>15,.2f}",
            f"{(member.tot_dividend_withdrawal if member.tot_dividend_withdrawal else 0.00):>15,.2f}",
            f"{(member.tot_loans if member.tot_loans else 0.00):>15,.2f}",
            f"{(member.tot_guaranteed if member.tot_guaranteed else 0.00):>15,.2f}",
            
            f"{(member.tot_guaranted if member.tot_guaranted else 0.00):>15,.2f}",
            f"{(member.available_balance if member.available_balance else 0.00):>15,.2f}",
            f"{(member.tot_interest_accrued if member.tot_interest_accrued else 0.00):>15,.2f}",
            f"{(member.tot_sav_int_deferred if member.tot_sav_int_deferred else 0.00):>15,.2f}",
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 1), (6, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)

    # Summary
    total_deposits = sum(m.tot_deposits or 0 for m in members)
    tot_dep_withdraw = sum(m.tot_deposit_withdrawal or 0 for m in members)
    total_shares = sum(m.tot_shares or 0 for m in members)
    total_shares_withdraw = sum(m.tot_shares_withdrawal or 0 for m in members)
    total_dividend = sum(m.tot_dividend or 0 for m in members)
    total_dividend_withdraw = sum(m.tot_dividend_withdrawal or 0 for m in members)
    total_loans = sum(m.tot_loans or 0 for m in members)
    total_guaranteed = sum(m.tot_guaranteed or 0 for m in members)
    total_guaranted = sum(m.tot_guaranted or 0 for m in members)
    total_avail_bal = sum(m.available_balance or 0 for m in members)
    total_interest_accrued = sum(m.tot_interest_accrued for m in members)
    total_interest_deferred = sum(m.tot_sav_int_deferred for m in members)

    elements.append(Spacer(1, 0.5*cm))
    summary_data = [
        ["SUMMARY TOTALS", "AMOUNT(GHS)"],
        ["Total Deposits:", f"{total_deposits:,.2f}"],
        ["Total Deposits Withdrawn:", f"{tot_dep_withdraw:,.2f}"],
        ["Total Shares:", f"{total_shares:,.2f}"],
        ["Total Shares Withdrawn:", f"{total_shares_withdraw:,.2f}"],
        ["Total Dividends:", f"{total_dividend:,.2f}"],
        ["Total Dividends Withdraw:", f"{total_dividend_withdraw:,.2f}"],
        ["Total Loans:", f"{total_loans:,.2f}"],
        ["Total Guaranteed:", f"{total_guaranteed:,.2f}"],
        ["Total Guaranted:", f" {total_guaranted:,.2f}"],
        ["Total Available Balance:", f" {total_avail_bal:,.2f}"],
        ["Total Interest Accrued:", f" {total_interest_accrued:,.2f}"],
        ["Total Interest Deferred:", f" {total_interest_deferred:,.2f}"],
    
    ]

    summary_table = Table(summary_data, colWidths=[6*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(summary_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Financial_Report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


# ##======================================== Excel Reports ====================================
# MembersApp/views_reports_excel.py
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from .models import Master
from django_ledger.models import EntityModel


def get_excel_style():
    """Return plain Excel styles"""
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return header_font, header_alignment, cell_alignment, number_alignment, thin_border


@login_required
def members_info_excel(request, slug):
    """Excel Report - Members basic information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Information"

    header_font, header_alignment, cell_alignment, number_alignment, thin_border = (
        get_excel_style()
    )

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "MEMBERS INFORMATION REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "ID",
        "Full Name",
        "Title",
        "Gender",
        "Date of Birth",
        "Date Enrolled",
        "Status",
        "Role",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    members = Master.objects.filter(is_deleted=False).order_by(
        "last_name", "first_name"
    )

    for member in members:
        ws.append(
            [
                member.id,
                member.full_name or "-",
                member.get_title_display(),
                member.get_gender_display(),
                (
                    member.date_of_birth.strftime("%Y-%m-%d")
                    if member.date_of_birth
                    else "-"
                ),
                (
                    member.date_enrolled.strftime("%Y-%m-%d")
                    if member.date_enrolled
                    else "-"
                ),
                member.get_mem_status_display(),
                member.get_role_display(),
            ]
        )

        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Members_Information_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def members_contact_excel(request, slug):
    """Excel Report - Members contact information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Contact"

    header_font, header_alignment, cell_alignment, number_alignment, thin_border = (
        get_excel_style()
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "MEMBERS CONTACT REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "ID",
        "Full Name",
        "Phone 1",
        "Phone 2",
        "Email",
        "City",
        "Postal Address",
        "Residential Address",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    members = Master.objects.filter(is_deleted=False).order_by(
        "last_name", "first_name"
    )

    for member in members:
        ws.append(
            [
                member.id,
                member.full_name or "-",
                member.telephone1 or "-",
                member.telephone2 or "-",
                member.email_address or "-",
                member.city or "-",
                member.postal_address or "-",
                member.residential_address or "-",
            ]
        )

        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 30
    ws.column_dimensions["H"].width = 30

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Members_Contact_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def next_of_kin_excel(request, slug):
    """Excel Report - Next of Kin information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    wb = Workbook()
    ws = wb.active
    ws.title = "Next of Kin"

    header_font, header_alignment, cell_alignment, number_alignment, thin_border = (
        get_excel_style()
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "NEXT OF KIN REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "ID",
        "Member Name",
        "NOK Name",
        "NOK Address",
        "NOK Phone",
        "Relation",
        "Percentage",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    members = Master.objects.filter(is_deleted=False).order_by(
        "last_name", "first_name"
    )

    for member in members:
        if member.nok_name1:
            ws.append(
                [
                    member.id,
                    member.full_name or "-",
                    member.nok_name1,
                    member.nok_address1 or "-",
                    member.nok_telephone1 or "-",
                    member.nok_relation1 or "-",
                    f"{member.nok_percent1 or 0}%",
                ]
            )
            for col in range(1, 8):
                ws.cell(row=ws.max_row, column=col).border = thin_border

        if member.nok_name2:
            ws.append(
                [
                    member.id,
                    member.full_name or "-",
                    member.nok_name2,
                    member.nok_address2 or "-",
                    member.nok_telephone2 or "-",
                    member.nok_relation2 or "-",
                    f"{member.nok_percent2 or 0}%",
                ]
            )
            for col in range(1, 8):
                ws.cell(row=ws.max_row, column=col).border = thin_border

        if member.nok_name3:
            ws.append(
                [
                    member.id,
                    member.full_name or "-",
                    member.nok_name3,
                    member.nok_address3 or "-",
                    member.nok_telephone3 or "-",
                    member.nok_relation3 or "-",
                    f"{member.nok_percent3 or 0}%",
                ]
            )
            for col in range(1, 8):
                ws.cell(row=ws.max_row, column=col).border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 10

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Next_Of_Kin_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def financial_report_excel(request, slug):
    """Excel Report - Financial information"""
    entity = get_object_or_404(EntityModel, slug=slug)
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    header_font, header_alignment, cell_alignment, number_alignment, thin_border = (
        get_excel_style()
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "MEMBERS FINANCIAL REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "ID",
        "Full Name",
        "Total Deposits",
        "Total Shares",
        "Total Interest",
        "Total Dividend",
        "Available Balance",
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    members = Master.objects.filter(is_deleted=False).order_by(
        "last_name", "first_name"
    )
    total_deposits = 0
    total_shares = 0
    total_balance = 0

    for member in members:
        deposits = member.tot_deposits or 0
        shares = member.tot_shares or 0
        balance = member.available_balance

        total_deposits += deposits
        total_shares += shares
        total_balance += balance

        ws.append(
            [
                member.id,
                member.full_name or "-",
                deposits,
                shares,
                member.tot_interest_accrued or 0,
                member.tot_dividend or 0,
                balance,
            ]
        )

        for col in range(1, 8):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.border = thin_border
            if col >= 3:
                cell.alignment = number_alignment
                cell.number_format = "#,##0.00"

    # Add totals row
    row = ws.max_row + 2
    ws[f"A{row}"] = "TOTALS"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"C{row}"] = total_deposits
    ws[f"D{row}"] = total_shares
    ws[f"G{row}"] = total_balance
    for col in ["C", "D", "G"]:
        ws[f"{col}{row}"].number_format = "#,##0.00"
        ws[f"{col}{row}"].alignment = number_alignment

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 15

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Financial_Report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)
    return response
