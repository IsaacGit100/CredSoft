# coa/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import IntegrityError

from .models import ChartOfAccounts



@login_required
def coa_home(request, slug):
    return render(request, 'coa/coa_home.html')

@login_required
def back_to_home(request):
    return redirect('/')
    

@login_required
def main_menu(request):
    return render(request, 'main_menu.html')


@login_required
def coa_create(request, slug):
    """Create a new account with duplicate checking."""
    
    if request.method == 'POST':
        account_type = request.POST.get('account_type')
        name = request.POST.get('name')
        behavior = request.POST.get('behavior', 'NORMAL')
        parent_id = request.POST.get('parent_account')
        is_data_entry = request.POST.get('is_data_entry') == 'on'
        is_data_view = request.POST.get('is_data_view', 'on') == 'on'
        
        parent = None
        if parent_id and parent_id != '':
            parent = get_object_or_404(ChartOfAccounts, id=parent_id)
            
            # Check if parent can have children
            if not parent.can_have_children():
                messages.error(
                    request, 
                    f"Cannot add sub-account to '{parent.name}' because it already has ledger entries."
                )
                return redirect('coa:coa_create')
        
        try:
            account = ChartOfAccounts(
                name=name,
                account_type=account_type,
                behavior=behavior,
                parent_account=parent,
                is_data_entry=is_data_entry,
                is_data_view=is_data_view,
            )
            account.save()
            
            messages.success(
                request, 
                f'Account "{name}" created successfully! Account No: {account.accountno}'
            )
            return redirect('coa:coa_detail', pk=account.pk)
            
        except IntegrityError as e:
            messages.error(request, f'Duplicate account number. Please try again.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    # GET request - show form
    parent_accounts = ChartOfAccounts.objects.filter(
        is_active=True,
        is_data_filled=False
    ).order_by('accountno')
    
    context = {
        'accounts': parent_accounts,
        'account_types': ChartOfAccounts.ACCOUNT_TYPES,
        'behavior_choices': ChartOfAccounts.BEHAVIOR_CHOICES,
    }
    return render(request, 'coa/coa_create.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models


# coa/views.py - Debug version
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from .models import ChartOfAccounts

@login_required
def coa_list(request):
    """List all accounts in hierarchical view."""
        
    # Get ALL accounts first (no filters)
    all_accounts = ChartOfAccounts.objects.all()
    print(f"Total accounts in DB: {all_accounts.count()}")
    
    for acc in all_accounts:
        print(f"  {acc.accountno} - {acc.name} (active: {acc.is_active}, parent: {acc.parent_account_id})")
    
    # Get root accounts (no parent) - DON'T filter by is_active yet
    root_accounts_list = ChartOfAccounts.objects.filter(parent_account__isnull=True)
    print(f"\nRoot accounts (no parent): {root_accounts_list.count()}")
    
    # Build flat list with levels
    def flatten_tree(account, level=0, result_list=None):
        if result_list is None:
            result_list = []
        account.level = level
        result_list.append(account)
        for child in account.children.all().order_by('accountno'):  # Get all children, don't filter
            flatten_tree(child, level + 1, result_list)
        return result_list
    
    # Create flat list with levels
    root_accounts = []
    for root in root_accounts_list:
        root_accounts.extend(flatten_tree(root, 0))
    
    
    # Calculate counts
    asset_count = ChartOfAccounts.objects.filter(account_type='ASSET').count()
    liability_count = ChartOfAccounts.objects.filter(account_type='LIABILITY').count()
    equity_count = ChartOfAccounts.objects.filter(account_type='EQUITY').count()
    income_count = ChartOfAccounts.objects.filter(account_type='INCOME').count()
    expense_count = ChartOfAccounts.objects.filter(account_type='EXPENSE').count()
    
    
    context = {
        'root_accounts': root_accounts,
        'total_accounts': all_accounts.count(),
        'asset_count': asset_count,
        'liability_count': liability_count,
        'equity_count': equity_count,
        'income_count': income_count,
        'expense_count': expense_count,
    }
    return render(request, 'coa/coa_list.html', context)




@login_required
def coa_detail(request, pk):
    """View account details."""
    account = get_object_or_404(ChartOfAccounts, pk=pk)
    children = ChartOfAccounts.objects.filter(parent_account=account, is_active=True)
    parent = account.parent_account
    
    context = {
        'account': account,
        'children': children,
        'parent': parent,
        'can_add_children': account.can_have_children(),
    }
    return render(request, 'coa/coa_detail.html', context)

@login_required
def coa_edit(request, pk):
    """Edit account."""
    account = get_object_or_404(ChartOfAccounts, pk=pk)
    
    if request.method == 'POST':
        account.name = request.POST.get('name')
        account.behavior = request.POST.get('behavior')
        account.is_data_entry = request.POST.get('is_data_entry') == 'on'
        account.is_data_view = request.POST.get('is_data_view') == 'on'
        account.save()
        
        messages.success(request, f'Account "{account.name}" updated successfully!')
        return redirect('coa:coa_detail', pk=account.pk)
    
    context = {
        'account': account,
        'behavior_choices': ChartOfAccounts.BEHAVIOR_CHOICES,
    }
    return render(request, 'coa/coa_edit.html', context)

@login_required
def coa_delete(request, pk):
    """Delete an account (only if no children)."""
    account = get_object_or_404(ChartOfAccounts, pk=pk)
    
    if request.method == 'POST':
        if account.children.exists():
            messages.error(request, f'Cannot delete "{account.name}" because it has sub-accounts.')
        elif account.is_data_filled:
            messages.error(request, f'Cannot delete "{account.name}" because it has ledger entries.')
        else:
            account.delete()
            messages.success(request, f'Account "{account.name}" deleted successfully!')
            return redirect('coa:coa_list')
    
    context = {'account': account}
    return render(request, 'coa/coa_confirm_delete.html', context)


# coa/views.py - Add these functions

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db import connection
from django.http import JsonResponse
from .models import ChartOfAccounts

@login_required
@permission_required('coa.delete_chartofaccounts', raise_exception=True)
def coa_reset(request):
    """Reset Chart of Accounts - Delete all accounts"""
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        
        if confirm == 'YES':
            try:
                # Count before deletion
                count = ChartOfAccounts.objects.count()
                
                # Disable foreign key checks
                with connection.cursor() as cursor:
                    cursor.execute("SET FOREIGN_KEY_CHECKS = 0;")
                    
                    # Delete all accounts
                    ChartOfAccounts.objects.all().delete()
                    
                    # Reset auto-increment
                    cursor.execute("ALTER TABLE coa_chartofaccounts AUTO_INCREMENT = 1;")
                    
                    # Re-enable foreign key checks
                    cursor.execute("SET FOREIGN_KEY_CHECKS = 1;")
                
                messages.success(request, f'✅ Chart of Accounts reset successfully! Deleted {count} account(s).')
                
                # Check if we should initialize default accounts
                if request.POST.get('initialize') == 'on':
                    initialize_default_accounts(request)
                    messages.success(request, 'Default accounts initialized!')
                
                return redirect('coa:coa_list')
                
            except Exception as e:
                messages.error(request, f'Error resetting Chart of Accounts: {str(e)}')
                return redirect('coa:coa_reset')
        else:
            messages.error(request, 'Reset cancelled. Please type "YES" to confirm.')
            return redirect('coa:coa_reset')
    
    # GET request - show confirmation page
    context = {
        'total_accounts': ChartOfAccounts.objects.count(),
        'root_accounts': ChartOfAccounts.objects.filter(parent_account__isnull=True),
    }
    return render(request, 'coa/coa_reset.html', context)


def initialize_default_accounts(request):
    """Initialize default Chart of Accounts"""
    
    # ============================================================
    # LEVEL 1: Major Categories
    # ============================================================
    assets = ChartOfAccounts.objects.create(
        name='ASSETS',
        account_type='ASSET',
        behavior='NORMAL',
        is_data_entry=False,
    )
    
    liabilities = ChartOfAccounts.objects.create(
        name='LIABILITIES',
        account_type='LIABILITY',
        behavior='NORMAL',
        is_data_entry=False,
    )
    
    equity = ChartOfAccounts.objects.create(
        name="EQUITY",
        account_type='EQUITY',
        behavior='NORMAL',
        is_data_entry=False,
    )
    
    income = ChartOfAccounts.objects.create(
        name='INCOME',
        account_type='INCOME',
        behavior='NORMAL',
        is_data_entry=False,
    )
    
    expenses = ChartOfAccounts.objects.create(
        name='EXPENSES',
        account_type='EXPENSE',
        behavior='NORMAL',
        is_data_entry=False,
    )
    
    # ============================================================
    # LEVEL 2: Sub Categories under ASSETS
    # ============================================================
    assets = ChartOfAccounts.objects.get(accountno='10000000')
    
    current_assets = ChartOfAccounts.objects.create(
        name='CURRENT ASSETS',
        account_type='ASSET',
        behavior='NORMAL',
        parent_account=assets,
        is_data_entry=False,
    )
    
    # ============================================================
    # LEVEL 3: Groups under CURRENT ASSETS
    # ============================================================
    current_assets = ChartOfAccounts.objects.get(accountno='10100000')
    
    cash = ChartOfAccounts.objects.create(
        name='CASH AND CASH EQUIVALENTS',
        account_type='ASSET',
        behavior='NORMAL',
        parent_account=current_assets,
        is_data_entry=False,
    )
    
    bank = ChartOfAccounts.objects.create(
        name='BANK ACCOUNTS',
        account_type='ASSET',
        behavior='NORMAL',
        parent_account=current_assets,
        is_data_entry=False,
    )
    
    # ============================================================
    # LEVEL 4: Specific Accounts
    # ============================================================
    cash = ChartOfAccounts.objects.get(accountno='10101000')
    
    ChartOfAccounts.objects.create(
        name='Cash in Hand - Main Office',
        account_type='ASSET',
        behavior='CASH',
        parent_account=cash,
        is_data_entry=True,
    )
    
    ChartOfAccounts.objects.create(
        name='Cash in Hand - Branch Office',
        account_type='ASSET',
        behavior='CASH',
        parent_account=cash,
        is_data_entry=True,
    )
    
    # Bank accounts
    bank = ChartOfAccounts.objects.get(accountno='10102000')
    
    ChartOfAccounts.objects.create(
        name='GCB Bank PLC',
        account_type='ASSET',
        behavior='BANK',
        parent_account=bank,
        is_data_entry=True,
    )
    
    ChartOfAccounts.objects.create(
        name='Absa Bank PLC',
        account_type='ASSET',
        behavior='BANK',
        parent_account=bank,
        is_data_entry=True,
    )
    
    # ============================================================
    # LEVEL 2: Sub Categories under LIABILITIES
    # ============================================================
    liabilities = ChartOfAccounts.objects.get(accountno='20000000')
    
    member_funds = ChartOfAccounts.objects.create(
        name='MEMBER FUNDS',
        account_type='LIABILITY',
        behavior='NORMAL',
        parent_account=liabilities,
        is_data_entry=False,
    )
    
    # Level 3
    member_funds = ChartOfAccounts.objects.get(accountno='20100000')
    
    savings = ChartOfAccounts.objects.create(
        name='SAVINGS ACCOUNTS',
        account_type='LIABILITY',
        behavior='SAVINGS',
        parent_account=member_funds,
        is_data_entry=False,
    )
    
    # Level 4
    savings = ChartOfAccounts.objects.get(accountno='20101000')
    
    ChartOfAccounts.objects.create(
        name='Regular Savings Account',
        account_type='LIABILITY',
        behavior='SAVINGS',
        parent_account=savings,
        is_data_entry=True,
    )
    
    ChartOfAccounts.objects.create(
        name='Target Savings Account',
        account_type='LIABILITY',
        behavior='SAVINGS',
        parent_account=savings,
        is_data_entry=True,
    )
    
    return True


# coa/views.py - Add this view
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.management import call_command
from io import StringIO

@login_required
@user_passes_test(lambda u: u.is_superuser)
def init_coa_from_web(request):
    """Initialize Chart of Accounts from web interface"""
    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        if confirm == 'YES':
            try:
                # Capture command output
                out = StringIO()
                call_command('init_full_coa', force=True, stdout=out)
                output = out.getvalue()
                
                messages.success(request, "Chart of Accounts initialized successfully!")
                return redirect('coa:coa_list')
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            messages.error(request, "Please type 'YES' to confirm")
            return redirect('coa:init_coa')
    
    return render(request, 'coa/init_coa_confirm.html')

# coa/views.py - Add this function
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.management import call_command
from io import StringIO

@login_required
def coa_init_standalone(request):
    """Standalone page to initialize Chart of Accounts"""
    
    if request.method == 'POST':
        confirm = request.POST.get('confirm')
        
        if confirm == 'YES':
            try:
                # Capture command output
                out = StringIO()
                call_command('init_full_coa', force=True, stdout=out)
                output = out.getvalue()
                
                messages.success(request, "✅ Chart of Accounts initialized successfully!")
                return redirect('coa:coa_list')
                
            except Exception as e:
                messages.error(request, f"❌ Error: {str(e)}")
                return redirect('coa:coa_init_standalone')
        else:
            messages.error(request, "Please type 'YES' to confirm")
            return redirect('coa:coa_init_standalone')
    
    # GET request - show the standalone page
    return render(request, 'coa/coa_init_standalone.html')




# coa/views.py - Add this function
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from .models import ChartOfAccounts

@login_required
def coa_pdf(request):
    """Generate PDF report of Chart of Accounts"""
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create document in landscape for more columns
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.gray,
        spaceAfter=30
    )
    
    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.white,
        alignment=TA_CENTER
    )
    
    # Get all root accounts (top level)
    root_accounts = ChartOfAccounts.objects.filter(
        parent_account__isnull=True,
        is_active=True
    ).order_by('accountno')
    
    # Build elements list
    elements = []
    
    # Title
    title = Paragraph("ST ANDREWS CO-OPERATIVE CREDIT UNION", title_style)
    elements.append(title)
    
    subtitle = Paragraph("Chart of Accounts", subtitle_style)
    elements.append(subtitle)
    
    # Generation date
    from datetime import datetime
    date_para = Paragraph(
        f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    )
    elements.append(date_para)
    elements.append(Spacer(1, 0.2*inch))
    
    # Build hierarchical table data
    def build_account_rows(account, level=0, rows=None):
        if rows is None:
            rows = []
        
        # Format account number with hyphens for display
        if len(account.accountno) == 8:
            formatted_no = f"{account.accountno[0]}-{account.accountno[1:3]}-{account.accountno[3:5]}-{account.accountno[5:8]}"
        else:
            formatted_no = account.accountno
        
        # Account type badge
        type_badge = account.get_account_type_display()
        
        # Behavior badge
        behavior_badge = account.get_behavior_display() if account.behavior != 'NORMAL' else '-'
        
        # Data entry indicator
        data_entry = "✓" if account.is_data_entry else ""
        
        # Indentation for hierarchy
        indent = "  " * level
        name_display = f"{indent}{account.name}"
        
        rows.append({
            'level': level,
            'code': formatted_no,
            'name': name_display,
            'type': type_badge,
            'behavior': behavior_badge,
            'data_entry': data_entry,
            'status': "Active" if account.is_active else "Inactive"
        })
        
        # Add children
        for child in account.children.filter(is_active=True).order_by('accountno'):
            build_account_rows(child, level + 1, rows)
        
        return rows
    
    # Build all rows
    all_rows = []
    for root in root_accounts:
        all_rows.extend(build_account_rows(root, 0))
    
    # Prepare table data
    table_data = []
    
    # Header row
    table_data.append([
        Paragraph("<b>Account No</b>", heading_style),
        Paragraph("<b>Account Name</b>", heading_style),
        Paragraph("<b>Type</b>", heading_style),
        Paragraph("<b>Behavior</b>", heading_style),
        Paragraph("<b>Data Entry</b>", heading_style),
        Paragraph("<b>Status</b>", heading_style),
    ])
    
    # Data rows
    for row in all_rows:
        # Style based on level
        name_style = ParagraphStyle(
            'LevelStyle',
            parent=styles['Normal'],
            leftIndent=row['level'] * 15
        )
        
        table_data.append([
            Paragraph(f"<font size='9'>{row['code']}</font>", styles['Normal']),
            Paragraph(f"<font size='9'>{row['name']}</font>", name_style),
            Paragraph(f"<font size='9'>{row['type']}</font>", styles['Normal']),
            Paragraph(f"<font size='9'>{row['behavior']}</font>", styles['Normal']),
            Paragraph(f"<font size='9'>{row['data_entry']}</font>", styles['Normal']),
            Paragraph(f"<font size='9'>{row['status']}</font>", styles['Normal']),
        ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Table styling
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Row styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#2C3E50')),
        
        # Column widths
        ('COLWIDTHS', (0, 0), (0, 0), 1.2*inch),
        ('COLWIDTHS', (1, 1), (1, 1), 3*inch),
        ('COLWIDTHS', (2, 2), (2, 2), 1*inch),
        ('COLWIDTHS', (3, 3), (3, 3), 1*inch),
        ('COLWIDTHS', (4, 4), (4, 4), 0.8*inch),
        ('COLWIDTHS', (5, 5), (5, 5), 0.8*inch),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Footer with summary
    total_accounts = ChartOfAccounts.objects.filter(is_active=True).count()
    data_entry_count = ChartOfAccounts.objects.filter(is_active=True, is_data_entry=True).count()
    
    summary_text = f"""
    <para align="center" fontSize="9">
    <b>Summary:</b> Total Accounts: {total_accounts} | Data Entry Accounts: {data_entry_count}
    </para>
    """
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Create response
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Chart_of_Accounts_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    return response


# coa/views_excel.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal
from .models import ChartOfAccounts
from FinanceApp.models import GeneralLedger

@login_required
def coa_excel(request):
    """Export Chart of Accounts to Excel"""
    
    # Create workbook
    wb = Workbook()
    
    # ============================================================
    # SHEET 1: Complete Chart of Accounts
    # ============================================================
    ws1 = wb.active
    ws1.title = "Chart of Accounts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    subheader_font_color = Font(bold=True, color="FFFFFF", size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value = "CHART OF ACCOUNTS"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    ws1.merge_cells('A2:H2')
    subtitle_cell = ws1['A2']
    subtitle_cell.value = "St Andrews Co-Operative Credit Union"
    subtitle_cell.font = Font(size=12)
    subtitle_cell.alignment = Alignment(horizontal="center")
    
    from datetime import datetime
    ws1.merge_cells('A3:H3')
    date_cell = ws1['A3']
    date_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    date_cell.font = Font(size=10, italic=True)
    date_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['Account Code', 'Formatted Code', 'Account Name', 'Account Type', 
               'Behavior', 'Parent Account', 'Data Entry', 'Current Balance']
    ws1.append([''] * 8)  # Empty row
    ws1.append(headers)
    
    # Style headers
    for col in range(1, 9):
        cell = ws1.cell(row=5, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get all active accounts
    accounts = ChartOfAccounts.objects.filter(is_active=True).order_by('accountno')
    
    # Function to format account number
    def format_account_no(accountno):
        if len(accountno) == 8:
            return f"{accountno[0]}-{accountno[1:3]}-{accountno[3:5]}-{accountno[5:8]}"
        return accountno
    
    # Get parent name
    def get_parent_name(account):
        if account.parent_account:
            return f"{account.parent_account.accountno} - {account.parent_account.name}"
        return "Top Level"
    
    # Get balance from ledger
    def get_account_balance(account):
        try:
            ledger = GeneralLedger.objects.filter(account=account).first()
            if ledger:
                return float(ledger.current_balance)
            return 0.00
        except:
            return 0.00
    
    # Add data rows
    for account in accounts:
        formatted_no = format_account_no(account.accountno)
        parent_name = get_parent_name(account)
        balance = get_account_balance(account)
        
        ws1.append([
            account.accountno,
            formatted_no,
            account.name,
            account.get_account_type_display(),
            account.get_behavior_display(),
            parent_name,
            "Yes" if account.is_data_entry else "No",
            balance
        ])
        
        # Style data row
        row_num = ws1.max_row
        for col in range(1, 9):
            cell = ws1.cell(row=row_num, column=col)
            cell.border = thin_border
            if col == 1:  # Account code column
                cell.font = Font(bold=True)
            if col == 2:  # Formatted code column
                cell.font = Font(italic=True)
            if col == 8:  # Balance column
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # ============================================================
    # SHEET 2: Accounts by Type
    # ============================================================
    ws2 = wb.create_sheet("Accounts by Type")
    
    # Title
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "ACCOUNTS BY TYPE"
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    # Get account types
    account_types = ChartOfAccounts.ACCOUNT_TYPES
    row = 3
    
    for type_code, type_label in account_types:
        # Type header
        ws2.merge_cells(f'A{row}:D{row}')
        type_cell = ws2[f'A{row}']
        type_cell.value = type_label.upper()
        type_cell.font = Font(bold=True, size=12, color="FFFFFF")
        type_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        type_cell.alignment = Alignment(horizontal="center")
        row += 1
        
        # Sub headers
        ws2[f'A{row}'] = "Account Code"
        ws2[f'B{row}'] = "Account Name"
        ws2[f'C{row}'] = "Behavior"
        ws2[f'D{row}'] = "Balance"
        
        for col in range(1, 5):
            cell = ws2.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
        row += 1
        
        # Accounts of this type
        type_accounts = accounts.filter(account_type=type_code)
        for account in type_accounts:
            ws2[f'A{row}'] = account.accountno
            ws2[f'B{row}'] = account.name
            ws2[f'C{row}'] = account.get_behavior_display()
            ws2[f'D{row}'] = get_account_balance(account)
            ws2[f'D{row}'].number_format = '#,##0.00'
            row += 1
        
        row += 1  # Empty row between types
    
    # Adjust column widths for sheet 2
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 20
    
    # ============================================================
    # SHEET 3: Hierarchical View
    # ============================================================
    ws3 = wb.create_sheet("Hierarchical View")
    
    ws3.merge_cells('A1:C1')
    ws3['A1'] = "CHART OF ACCOUNTS - HIERARCHICAL VIEW"
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = Alignment(horizontal="center")
    
    ws3['A2'] = "Level"
    ws3['B2'] = "Account Code"
    ws3['C2'] = "Account Name"
    
    for col in range(1, 4):
        cell = ws3.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Build hierarchical tree
    def add_to_sheet(account, level, row):
        ws3.cell(row=row, column=1, value=level)
        ws3.cell(row=row, column=2, value=account.accountno)
        ws3.cell(row=row, column=3, value="  " * level + account.name)
        row += 1
        for child in account.children.filter(is_active=True).order_by('accountno'):
            row = add_to_sheet(child, level + 1, row)
        return row
    
    row = 3
    roots = ChartOfAccounts.objects.filter(parent_account__isnull=True, is_active=True).order_by('accountno')
    for root in roots:
        row = add_to_sheet(root, 0, row)
    
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 50
    
    # ============================================================
    # SHEET 4: Summary Statistics
    # ============================================================
    ws4 = wb.create_sheet("Summary Statistics")
    
    ws4.merge_cells('A1:B1')
    ws4['A1'] = "CHART OF ACCOUNTS SUMMARY"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4['A1'].alignment = Alignment(horizontal="center")
    
    # Statistics
    stats = [
        ("Total Accounts", accounts.count()),
        ("Active Accounts", accounts.filter(is_active=True).count()),
        ("Data Entry Accounts", accounts.filter(is_data_entry=True).count()),
        ("Data Filled Accounts", accounts.filter(is_data_filled=True).count()),
        ("Assets", accounts.filter(account_type='ASSET').count()),
        ("Liabilities", accounts.filter(account_type='LIABILITY').count()),
        ("Equity", accounts.filter(account_type='EQUITY').count()),
        ("Income", accounts.filter(account_type='INCOME').count()),
        ("Expenses", accounts.filter(account_type='EXPENSE').count()),
    ]
    
    row = 3
    for stat_name, stat_value in stats:
        ws4[f'A{row}'] = stat_name
        ws4[f'B{row}'] = stat_value
        ws4[f'B{row}'].alignment = Alignment(horizontal="right")
        row += 1
    
    # Balance totals
    row += 1
    ws4[f'A{row}'] = "BALANCE TOTALS"
    ws4[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Calculate totals by account type
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    total_income = 0
    total_expenses = 0
    
    for account in accounts:
        balance = get_account_balance(account)
        if account.account_type == 'ASSET':
            total_assets += balance
        elif account.account_type == 'LIABILITY':
            total_liabilities += balance
        elif account.account_type == 'EQUITY':
            total_equity += balance
        elif account.account_type == 'INCOME':
            total_income += balance
        elif account.account_type == 'EXPENSE':
            total_expenses += balance
    
    balance_stats = [
        ("Total Assets", total_assets),
        ("Total Liabilities", total_liabilities),
        ("Total Equity", total_equity),
        ("Total Income", total_income),
        ("Total Expenses", total_expenses),
        ("Net Position", total_assets - (total_liabilities + total_equity)),
    ]
    
    for stat_name, stat_value in balance_stats:
        ws4[f'A{row}'] = stat_name
        ws4[f'B{row}'] = stat_value
        ws4[f'B{row}'].number_format = '#,##0.00'
        ws4[f'B{row}'].alignment = Alignment(horizontal="right")
        row += 1
    
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 20
    
    # ============================================================
    # SAVE RESPONSE
    # ============================================================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Chart_of_Accounts_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response

