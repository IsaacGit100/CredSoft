from django.shortcuts import render

# Create your views here.


def fixed_assets_home(request):
    return render(request, 'FixedAssets/fixed_assets_home.html')

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import AssetCategory, FixedAsset, DepreciationEntry
from coa.models import ChartOfAccounts
from .forms import AssetCategoryForm, FixedAssetForm


@login_required
@staff_member_required
def category_list_manage(request):
    categories = AssetCategory.objects.all().order_by('name')
    return render(request, 'FixedAssets/category_list_manage.html', {'categories': categories})

@login_required
@staff_member_required
def category_create(request):
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset category created.')
            return redirect('FixedAssets:category_list_manage')
    else:
        form = AssetCategoryForm()
    return render(request, 'FixedAssets/category_form.html', {'form': form})

@login_required
@staff_member_required
def category_edit(request, pk):
    category = get_object_or_404(AssetCategory, pk=pk)
    if request.method == 'POST':
        form = AssetCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset category updated.')
            return redirect('FixedAssets:category_list')
    else:
        form = AssetCategoryForm(instance=category)
    return render(request, 'FixedAssets/category_form.html', {'form': form, 'category': category})

@login_required
@staff_member_required
def category_delete(request, pk):
    category = get_object_or_404(AssetCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Asset category deleted.')
        return redirect('FixedAssets:category_list_manage')
    return render(request, 'FixedAssets/category_confirm_delete.html', {'category': category})



@login_required
@staff_member_required
def assets_list_manage(request):
    assets = FixedAsset.objects.select_related('category').all().order_by('name')
    return render(request, 'FixedAssets/assets_list_manage.html', {'assets': assets})

@login_required
@staff_member_required
def asset_create(request):
    if request.method == 'POST':
        form = FixedAssetForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset added.')
            return redirect('FixedAssets:assets_list_manage')
    else:
        form = FixedAssetForm()
    return render(request, 'FixedAssets/asset_form.html', {'form': form})

@login_required
@staff_member_required
def asset_edit(request, pk):
    asset = get_object_or_404(FixedAsset, pk=pk)
    if request.method == 'POST':
        form = FixedAssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, 'Asset updated.')
            return redirect('FixedAssets:assets_list_manage')
    else:
        form = FixedAssetForm(instance=asset)
    return render(request, 'FixedAssets/asset_form.html', {'form': form, 'asset': asset})

@login_required
@staff_member_required
def asset_delete(request, pk):
    asset = get_object_or_404(FixedAsset, pk=pk)
    if request.method == 'POST':
        asset.delete()
        messages.success(request, 'Asset deleted.')
        return redirect('FixedAssets:asset_list_manage')
    return render(request, 'FixedAssets/asset_confirm_delete.html', {'asset': asset})


from decimal import Decimal
from django.db.models import Sum

@login_required
def asset_dashboard(request):
    assets = FixedAsset.objects.filter(is_active=True).select_related('category')
    total_cost = Decimal('0')
    total_depreciation = Decimal('0')
    total_nbv = Decimal('0')
    asset_data = []

    for asset in assets:
        acc_dep = asset.depreciation_entries.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        nbv = asset.cost - acc_dep
        total_cost += asset.cost
        total_depreciation += acc_dep
        total_nbv += nbv
        asset_data.append({
            'asset': asset,
            'acc_dep': acc_dep,
            'nbv': nbv,
        })

    context = {
        'asset_data': asset_data,
        'total_cost': total_cost,
        'total_depreciation': total_depreciation,
        'total_nbv': total_nbv,
        'asset_count': assets.count(),
    }
    return render(request, 'FixedAssets/dashboard.html', context)

@login_required
@staff_member_required
def post_depreciation(request):
    from django.utils import timezone
    from datetime import timedelta
    from FinanceApp.models import JournalEntry, JournalLine
    from decimal import Decimal

    if request.method == 'POST':
        # Process depreciation for all active assets
        posted = 0
        for asset in FixedAsset.objects.filter(is_active=True):
            # Calculate monthly depreciation (simple straight‑line)
            if asset.category.depreciation_method == 'SL':
                useful_months = asset.category.useful_life_years * 12
                monthly_dep = (asset.cost - asset.salvage_value) / useful_months
            else:
                # Reducing balance – simplified: use override or category rate
                rate = asset.override_depreciation_rate or 20.00
                monthly_rate = Decimal(rate) / 100 / 12
                nbv = asset.cost - asset.accumulated_depreciation
                monthly_dep = nbv * monthly_rate

            # Prevent over‑depreciation
            if (asset.cost - asset.accumulated_depreciation - monthly_dep) < asset.salvage_value:
                monthly_dep = asset.cost - asset.accumulated_depreciation - asset.salvage_value

            if monthly_dep <= 0:
                continue

            # Create journal entry
            today = timezone.now().date()
            journal = JournalEntry.objects.create(
                entry_number=f'DEP-{today.strftime("%Y%m")}-{asset.id}',
                entry_date=today,
                description=f'Depreciation for {asset.name}',
                status='POSTED',
                posted=True,
                posted_at=timezone.now()
            )
            JournalLine.objects.create(
                journal=journal,
                account=asset.category.depreciation_expense_account,
                debit=monthly_dep,
                credit=0,
                line_description=f'Depreciation expense - {asset.name}'
            )
            JournalLine.objects.create(
                journal=journal,
                account=asset.category.accumulated_depreciation_account,
                debit=0,
                credit=monthly_dep,
                line_description=f'Accumulated depreciation - {asset.name}'
            )
            # Record the entry
            DepreciationEntry.objects.create(
                asset=asset,
                period_start=today.replace(day=1),
                period_end=today,
                amount=monthly_dep,
                journal_entry=journal
            )
            posted += 1

        messages.success(request, f'Posted depreciation for {posted} assets.')
        return redirect('FixedAssets:dashboard')

    # GET – show confirmation and preview
    preview_data = []
    for asset in FixedAsset.objects.filter(is_active=True):
        # Calculate monthly depreciation (same logic)
        if asset.category.depreciation_method == 'SL':
            useful_months = asset.category.useful_life_years * 12
            if useful_months != 0:
                monthly_dep = (asset.cost - asset.salvage_value) / useful_months
            else:
                monthly_dep = 0
        else:
            rate = asset.override_depreciation_rate or 20.00
            monthly_rate = Decimal(rate) / 100 / 12
            nbv = asset.cost - asset.accumulated_depreciation
            monthly_dep = nbv * monthly_rate
        if monthly_dep > 0:
            preview_data.append({
                'asset': asset,
                'monthly_dep': monthly_dep,
            })
    context = {'preview': preview_data}
    return render(request, 'FixedAssets/post_depreciation.html', context)


def fixed_asset_register(request):
    assets = FixedAsset.objects.filter(is_active=True).select_related('category')
    data = []
    total_cost = Decimal('0')
    total_acc_dep = Decimal('0')
    total_nbv = Decimal('0')
    for asset in assets:
        acc_dep = asset.depreciation_entries.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        nbv = asset.cost - acc_dep
        data.append({
            'asset': asset,
            'acc_dep': acc_dep,
            'nbv': nbv,
        })
        total_cost += asset.cost
        total_acc_dep += acc_dep
        total_nbv += nbv
    context = {
        'data': data,
        'total_cost': total_cost,
        'total_acc_dep': total_acc_dep,
        'total_nbv': total_nbv,
    }
    return render(request, 'FixedAssets/fixed_assets_register.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from .models import DepreciationEntry, FixedAsset

@login_required
def depreciation_list_manage(request):
    # Get all depreciation entries with related asset
    entries = DepreciationEntry.objects.select_related('asset', 'journal_entry').all().order_by('-period_end')
    print(entries)
    
    # Calculate totals
    total_depreciation = entries.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Group by asset (optional summary)
    asset_summary = {}
    for entry in entries:
        asset_name = entry.asset.name
        if asset_name not in asset_summary:
            asset_summary[asset_name] = 0
        asset_summary[asset_name] += entry.amount
    
    context = {
        'entries': entries,
        'total_depreciation': total_depreciation,
        'asset_summary': asset_summary,
        'entry_count': entries.count(),
    }
    return render(request, 'FixedAssets/depreciation_list.html', context)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import DepreciationEntry

@login_required
def depreciation_export_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Schedule"

    # --- Headers ---
    headers = ['Asset ID', 'Asset Name', 'Period Start', 'Period End', 'Amount', 'Created At']
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Data ---
    entries = DepreciationEntry.objects.select_related('asset').all().order_by('-period_end')
    
    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.asset_id)
        ws.cell(row=row_idx, column=2, value=entry.asset.name)
        ws.cell(row=row_idx, column=3, value=entry.period_start.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=4, value=entry.period_end.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=5, value=float(entry.amount))
        ws.cell(row=row_idx, column=6, value=entry.created_at.strftime("%Y-%m-%d %H:%M"))

    # --- Format numbers ---
    for row in ws.iter_rows(min_row=2, max_row=len(entries)+1, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'

    # --- Total row ---
    total_row = len(entries) + 2
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=5, value=sum(e.amount for e in entries))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'

    # --- Auto-width ---
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # --- Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=depreciation_schedule.xlsx'
    wb.save(response)
    return response


from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from .models import DepreciationEntry

@login_required
def depreciation_export_pdf(request):
    entries = DepreciationEntry.objects.select_related('asset').all().order_by('-period_end')
    total_depreciation = sum(e.amount for e in entries)
    
    template = get_template('FixedAssets/depreciation_pdf.html')
    html = template.render({
        'entries': entries,
        'total_depreciation': total_depreciation,
    })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=depreciation_schedule.pdf'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF: ' + str(pisa_status.err))
    return response