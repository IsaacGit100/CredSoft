from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages

from .models import FixedAsset
from .forms import FixedAssetForm, AssetCategoryForm
from django.contrib.admin.views.decorators import staff_member_required
from .models import AssetCategory, FixedAsset, DepreciationEntry
from djan_led.utils import get_visible_accounts
from django.contrib.auth.models import User


from django.utils import timezone

from .models import FixedAsset, DepreciationEntry

from decimal import Decimal
from django.db.models import Sum


from .models import FixedAsset, DepreciationEntry
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from decimal import Decimal
from django_ledger.models import (
    EntityModel,
    LedgerModel,
    JournalEntryModel,
    AccountModel,
    TransactionModel,
)

from djan_led.utils import get_visible_accounts
from .forms import AssetCategoryForm
from .models import AssetCategory

import logging
logger = logging.getLogger(__name__)

@login_required
def fixed_assets_home(request, slug):
    return render(request, 'FixedAssets/fixed_assets_home.html')

@login_required
def asset_dashboard(request, slug):
    from django_ledger.models import EntityModel
    assets = FixedAsset.objects.filter(is_active=True).select_related('category')
    total_cost = Decimal('0')
    total_depreciation = Decimal('0')
    total_nbv = Decimal('0')
    asset_data = []

    for asset in assets:
        acc_dep = asset.depreciation_entries.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        nbv = asset.cost - acc_dep
        total_cost += asset.cost
        total_depreciation += acc_dep
        total_nbv += nbv
        asset_data.append({
            'asset': asset,
            'acc_dep': acc_dep,
            'nbv': nbv,
        })

    context = {
        'asset_data': asset_data,
        'total_cost': total_cost,
        'total_depreciation': total_depreciation,
        'total_nbv': total_nbv,
        'asset_count': assets.count(),
    }
    return render(request, 'FixedAssets/dashboard.html', context)


@login_required
def asset_list_manage(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)

    # Permission check
    try:
        profile = request.user.djan_led_profile
    except:
        profile = None

    if profile:
        if entity not in profile.allowed_entities.all() and entity != profile.default_entity:
            messages.error(request, "You do not have permission to access this entity.")
            return redirect('entity_dashboard', slug=profile.default_entity.slug)

    assets = FixedAsset.objects.filter(entity=entity, is_active=True).order_by('asset_id')

    context = {
        'entity': entity,
        'assets': assets,
    }
    return render(request, 'FixedAssets/asset_list_manage.html', context)


@login_required
def asset_add(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:fixed_assets_home", slug=slug)

    if request.method == "POST":
        print("🔥 POST received")  # Debug
        form = FixedAssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.entity = entity
            asset.save()
            messages.success(request, "Asset added.")
            return redirect("FixedAssets:asset_list_manage", slug=slug)
        else:
            print("Form errors:", form.errors)
    else:
        form = FixedAssetForm()

    return render(
        request, "FixedAssets/asset_form.html", {"form": form, "entity": entity}
    )


@login_required
def asset_add2(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:fixed_assets_home", slug=slug)

    # Get accounts (only if you need them for a dropdown – but we're not using them)
    # accounts = AccountModel.objects.filter(coa_model=coa, active=True, depth__gt=1).order_by("code")

    if request.method == "POST":
        form = FixedAssetForm(request.POST)
        #  REMOVE THIS LINE – no 'account' field in the form
        # form.fields["account"].queryset = accounts

        if form.is_valid():
            asset = form.save(commit=False)
            asset.entity = entity
            asset.save()
            messages.success(request, "Asset added.")
            return redirect("FixedAssets:fixed_assets_list", slug=slug)
    else:
        form = FixedAssetForm()
        #  REMOVE THIS LINE – no 'account' field in the form
        # form.fields["account"].queryset = accounts

    return render(
        request, "FixedAssets/asset_form.html", {"form": form, "entity": entity}
    )


def asset_add1(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    #    accounts = get_visible_accounts(request.user, entity)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("fixed_assets_home", slug=slug)

    # Get active, non‑root accounts for this entity
    accounts = AccountModel.objects.filter(
        coa_model=coa, active=True, depth__gt=1
    ).order_by("code")

    if request.method == "POST":
        form = FixedAssetForm(request.POST)
        # Set the account queryset on the form
        form.fields["account"].queryset = accounts
        if form.is_valid():
            asset = form.save(commit=False)
            asset.entity = entity
            asset.save()
            messages.success(request, "Asset added.")
            return redirect("fixed_assets_list", slug=slug)
    else:
        form = FixedAssetForm()
        form.fields["account"].queryset = accounts

    return render(
        request, "FixedAssets/asset_form.html", {"form": form, "entity": entity}
    )


@login_required
def asset_edit(request, slug, pk):
    entity = get_object_or_404(EntityModel, slug=slug)
    asset = get_object_or_404(FixedAsset, pk=pk, entity=entity)

    if request.method == "POST":
        form = FixedAssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, "Asset updated.")
            return redirect("FixedAssets:asset_list_manage", slug=slug)
    else:
        form = FixedAssetForm(instance=asset)

    return render(
        request,
        "FixedAssets/asset_form.html",
        {"form": form, "entity": entity, "asset": asset},
    )


@login_required
def asset_edit1(request, slug, asset_id):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    asset = get_object_or_404(FixedAsset, entity=entity, asset_id=asset_id)

    if request.method == 'POST':
        form = FixedAssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            messages.success(request, "Asset updated successfully.")
            return redirect('asset_list', slug=entity.slug)
    else:
        form = FixedAssetForm(instance=asset)

    context = {
        'entity': entity,
        'form': form,
        'asset': asset,
    }
    return render(request, 'FixedAssets/asset_form.html', context)


@login_required
def asset_delete(request, slug, pk):
    entity = get_object_or_404(EntityModel, slug=slug)
    asset = get_object_or_404(FixedAsset, pk=pk, entity=entity)

    if request.method == "POST":
        asset.delete()
        messages.success(request, "Asset deleted.")
        return redirect("FixedAssets:asset_list_manage", slug=slug)

    return render(
        request,
        "FixedAssets/asset_confirm_delete.html",
        {"asset": asset, "entity": entity},
    )


@login_required
def asset_delete1(request, slug, asset_id):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    asset = get_object_or_404(FixedAsset, entity=entity, asset_id=asset_id)

    if request.method == 'POST':
        asset.delete()
        messages.success(request, "Asset deleted successfully.")
        return redirect('asset_list', slug=entity.slug)

    context = {
        'entity': entity,
        'asset': asset,
    }
    return render(request, 'FixedAssets/asset_confirm_delete.html', context)


@login_required
def category_list_manage(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    categories = AssetCategory.objects.all().order_by('name')
    return render(request, 'FixedAssets/category_list_manage.html', {'categories': categories})


@login_required
def category_create(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:category_list_manage", slug=slug)

    # Get accounts (visible accounts for this user, or all active if no preference)
    accounts = get_visible_accounts(request.user, entity)
    # Fallback if get_visible_accounts is not defined:
    # accounts = AccountModel.objects.filter(coa_model=coa, active=True, depth__gt=1).order_by('code')

    if request.method == "POST":
        form = AssetCategoryForm(request.POST)

        # Assign the querysets to the form fields BEFORE validation
        form.fields["asset_account"].queryset = accounts
        form.fields["accumulated_depreciation_account"].queryset = accounts
        form.fields["depreciation_expense_account"].queryset = accounts

        if form.is_valid():
            category = form.save(commit=False)
            category.asset_account = form.cleaned_data["asset_account"]
            category.entity = entity  # ensure entity is set
            category.save()
            messages.success(request, "Asset category created.")
            return redirect("FixedAssets:category_list_manage", slug=slug)
    else:
        form = AssetCategoryForm()
        # Assign querysets for GET
        form.fields["asset_account"].queryset = accounts
        form.fields["accumulated_depreciation_account"].queryset = accounts
        form.fields["depreciation_expense_account"].queryset = accounts

    return render(
        request,
        "FixedAssets/category_form.html",
        {
            "form": form,
            "entity": entity,
            "category": None,
        },
    )


@login_required
def category_edit(request, pk, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:category_list_manage", slug=slug)

    # Get the category
    category = get_object_or_404(AssetCategory, pk=pk, entity=entity)

    # Get accounts (visible or all active)
    accounts = get_visible_accounts(request.user, entity)

    if request.method == "POST":
        form = AssetCategoryForm(request.POST, instance=category)
        #  Assign querysets before validation
        form.fields["asset_account"].queryset = accounts
        form.fields["accumulated_depreciation_account"].queryset = accounts
        form.fields["depreciation_expense_account"].queryset = accounts

        if form.is_valid():
            form.save()
            messages.success(request, "Asset category updated.")
            return redirect("FixedAssets:category_list_manage", slug=slug)
    else:
        form = AssetCategoryForm(instance=category)
        # Assign querysets for GET
        form.fields["asset_account"].queryset = accounts
        form.fields["accumulated_depreciation_account"].queryset = accounts
        form.fields["depreciation_expense_account"].queryset = accounts

    return render(
        request,
        "FixedAssets/category_form.html",
        {"form": form, "entity": entity, "category": category},
    )


@login_required
def category_delete(request, pk, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    category = get_object_or_404(AssetCategory, pk=pk, entity=entity)

    if request.method == "POST":
        # Check if any asset uses this category – optional
        if FixedAsset.objects.filter(category=category, is_active=True).exists():
            messages.error(
                request, "Cannot delete: This category is used by active assets."
            )
            return redirect("FixedAssets:category_list_manage", slug=slug)

        category.delete()
        messages.success(request, "Asset category deleted.")
        return redirect("FixedAssets:category_list_manage", slug=slug)

    return render(
        request,
        "FixedAssets/category_confirm_delete.html",
        {"category": category, "entity": entity},
    )


@login_required
@staff_member_required
def post_depreciation1(request, slug):
    """
    POST DEPRECIATION - Post monthly depreciation for all active assets of an entity.

    This is a **business service** that:
    1. Calculates depreciation for each active asset.
    2. Creates a Journal Entry in django-ledger for each asset.
    3. Posts the entry (debit Depreciation Expense, credit Accumulated Depreciation).
    4. Updates the asset's accumulated depreciation and book value.
    5. Records a DepreciationEntry for audit/history.

    The same pattern can be reused for any periodic posting (e.g., loan interest, savings interest, salary processing).
    """

    # ------------------------------------------------------------
    # 1. Get the entity (the organisation/church/credit union)
    # ------------------------------------------------------------
    entity = get_object_or_404(EntityModel, slug=slug)

    # Only process POST requests (triggered by a button)
    if request.method != "POST":
        return redirect("FixedAssets:asset_dashboard", slug=slug)

    # ------------------------------------------------------------
    # 2. Get or create a Ledger for this entity
    #    A ledger is a collection of journal entries (the "books").
    #    Each entity has its own ledger to keep data separate.
    # ------------------------------------------------------------
    ledger = LedgerModel.objects.filter(entity=entity).first()
    if not ledger:
        # If no ledger exists, create one with a default name.
        ledger = LedgerModel.objects.create(entity=entity, name="Default Ledger")

    # ------------------------------------------------------------
    # 3. Get the entity's Chart of Accounts (COA)
    #    The COA defines all the accounts (Cash, Bank, Revenue, Expenses, etc.)
    #    We need it to look up the specific GL accounts for depreciation.
    # ------------------------------------------------------------
    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:dashboard", slug=slug)

    # ------------------------------------------------------------
    # 4. Prepare counters and date ranges
    # ------------------------------------------------------------
    posted = 0
    today = timezone.now().date()
    month_start = today.replace(day=1)  # first day of current month

    # ------------------------------------------------------------
    # 5. Loop through all active assets of this entity
    # ------------------------------------------------------------
    for asset in FixedAsset.objects.filter(entity=entity, is_active=True):
        # Skip assets without a category (which defines the GL accounts)
        if not asset.category:
            continue

        # ==========================================================
        # 5a. CALCULATE MONTHLY DEPRECIATION
        # ==========================================================
        if asset.category.depreciation_method == "straight_line":
            # Straight‑line: (cost - salvage) / useful life in months
            useful_months = asset.category.useful_life_years * 12
            monthly_dep = (asset.cost - asset.salvage_value) / Decimal(useful_months)
        else:
            # Reducing balance (simplified):
            # monthly depreciation = net book value × monthly rate
            rate = Decimal(asset.override_depreciation_rate or 20.00)  # annual %
            monthly_rate = rate / 100 / 12
            nbv = asset.cost - asset.accumulated_depreciation
            monthly_dep = nbv * monthly_rate

        # Prevent over‑depreciation (asset cannot drop below salvage value)
        if (asset.cost - asset.accumulated_depreciation - monthly_dep) < asset.salvage_value:
            monthly_dep = asset.cost - asset.accumulated_depreciation - asset.salvage_value

        # If depreciation is zero or negative, skip this asset
        if monthly_dep <= Decimal("0.00"):
            continue

        # ==========================================================
        # 5b. GET THE GL ACCOUNTS FROM THE CHART OF ACCOUNTS
        #     These accounts are defined in the AssetCategory.
        # ==========================================================
        try:
            # Depreciation Expense Account (e.g., 6060)
            expense_account = AccountModel.objects.get(
                coa_model=coa,
                code=asset.category.depreciation_expense_account.code
            )
            # Accumulated Depreciation Account (e.g., 1099)
            acc_dep_account = AccountModel.objects.get(
                coa_model=coa,
                code=asset.category.accumulated_depreciation_account.code
            )
        except AccountModel.DoesNotExist:
            # If the account is missing, skip this asset and show an error
            messages.error(
                request,
                f"Account not found for {asset.name}. Check your category accounts.",
            )
            continue

        # ==========================================================
        # 5c. CREATE A JOURNAL ENTRY (DRAFT)
        #     A Journal Entry is a single accounting transaction.
        #     It must have at least one debit and one credit.
        # ==========================================================
        je = JournalEntryModel.objects.create(
            ledger=ledger,                 # Which ledger (entity) this belongs to
            timestamp=today,               # Date of the entry
            description=f"Depreciation for {asset.name}",
            posted=False,                  # Draft – we'll post it after adding lines
        )

        # ==========================================================
        # 5d. ADD TRANSACTIONS (DEBIT AND CREDIT)
        #     TransactionModel represents a single line in a journal entry.
        #     - Debit: Depreciation Expense (increases expense)
        #     - Credit: Accumulated Depreciation (contra‑asset, reduces asset value)
        # ==========================================================
        # Debit Depreciation Expense
        TransactionModel.objects.create(
            journal_entry=je,
            account=expense_account,
            amount=monthly_dep,
            tx_type="debit",
        )

        # Credit Accumulated Depreciation
        TransactionModel.objects.create(
            journal_entry=je,
            account=acc_dep_account,
            amount=monthly_dep,
            tx_type="credit",
        )

        # ==========================================================
        # 5e. POST THE JOURNAL ENTRY
        #     Posting validates that total debits = total credits.
        #     Once posted, the entry becomes permanent and affects the GL.
        # ==========================================================
        je.posted = True
        je.save()

        # ==========================================================
        # 5f. UPDATE THE ASSET RECORD
        #     Accumulated depreciation and book value are stored on the asset.
        # ==========================================================
        asset.accumulated_depreciation = (asset.accumulated_depreciation or Decimal("0")) + monthly_dep
        asset.book_value = asset.cost - asset.accumulated_depreciation
        asset.last_depreciation_date = today
        asset.save()

        # ==========================================================
        # 5g. RECORD A DEPRECIATION ENTRY (for audit/history)
        #     This keeps a historical record of each depreciation run.
        # ==========================================================
        DepreciationEntry.objects.create(
            asset=asset,
            period_start=month_start,
            period_end=today,
            amount=monthly_dep,
            journal_entry=je,
            created_by=request.user,
        )

        # Increment the counter
        posted += 1

    # ------------------------------------------------------------
    # 6. Return a success message and redirect
    # ------------------------------------------------------------
    messages.success(request, f"Posted depreciation for {posted} assets.")
    return redirect("FixedAssets:dashboard", slug=slug)


from django.db import transaction


@login_required
@staff_member_required
def post_depreciation(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)

    if request.method != "POST":
        return redirect("FixedAssets:asset_dashboard", slug=slug)

    ledger = LedgerModel.objects.filter(entity=entity).first()
    if not ledger:
        ledger = LedgerModel.objects.create(entity=entity, name="Default Ledger")

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:asset_dashboard", slug=slug)

    assets = FixedAsset.objects.filter(entity=entity, is_active=True)
    posted_assets = []
    total_amount = Decimal("0.00")
    today = timezone.now()
    month_start = today.replace(day=1)

    for asset in assets:
        if not asset.category:
            continue

        if (
            not asset.category.depreciation_expense_account
            or not asset.category.accumulated_depreciation_account
        ):
            continue

        # Calculate monthly depreciation
        if asset.category.depreciation_method == "straight_line":
            useful_months = asset.category.useful_life_years * 12
            monthly_dep = (asset.cost - asset.salvage_value) / Decimal(useful_months)
        else:
            rate = Decimal(
                asset.override_depreciation_rate
                or asset.category.depreciation_rate
                or 20.00
            )
            monthly_rate = rate / 100 / 12
            nbv = asset.cost - asset.accumulated_depreciation
            monthly_dep = nbv * monthly_rate

        if (
            asset.cost - asset.accumulated_depreciation - monthly_dep
        ) < asset.salvage_value:
            monthly_dep = (
                asset.cost - asset.accumulated_depreciation - asset.salvage_value
            )

        if monthly_dep <= Decimal("0.00"):
            continue

        # Get accounts
        try:
            expense_account = AccountModel.objects.get(
                coa_model=coa, code=asset.category.depreciation_expense_account.code
            )
            acc_dep_account = AccountModel.objects.get(
                coa_model=coa, code=asset.category.accumulated_depreciation_account.code
            )
        except AccountModel.DoesNotExist:
            continue

        # Create and post journal entry
        je = JournalEntryModel.objects.create(
            ledger=ledger,
            timestamp=today,
            description=f"Depreciation for {asset.name}",
            posted=False,
        )

        TransactionModel.objects.create(
            journal_entry=je,
            account=expense_account,
            amount=monthly_dep,
            tx_type="debit",
        )
        TransactionModel.objects.create(
            journal_entry=je,
            account=acc_dep_account,
            amount=monthly_dep,
            tx_type="credit",
        )

        je.posted = True
        je.save()

        # Update asset
        asset.accumulated_depreciation = (
            asset.accumulated_depreciation or Decimal("0")
        ) + monthly_dep
        asset.book_value = asset.cost - asset.accumulated_depreciation
        asset.last_depreciation_date = today
        asset.save()

        # Record depreciation entry
        DepreciationEntry.objects.create(
            asset=asset,
            period_start=month_start,
            period_end=today,
            amount=monthly_dep,
            journal_entry=je,
        )

        posted_assets.append(
            {
                "asset": asset,
                "amount": monthly_dep,
                "journal_entry": je,
            }
        )
        total_amount += monthly_dep

    context = {
        "entity": entity,
        "posted_count": len(posted_assets),
        "posted_assets": posted_assets,
        "total_amount": total_amount,
        "today": timezone.now()
    }
    return render(request, "FixedAssets/depreciation_success.html", context)


@login_required
@staff_member_required
def post_depreciation2(request, slug):

    entity = get_object_or_404(EntityModel, slug=slug)
    if request.method != "POST":
        return redirect("FixedAssets:asset_dashboard", slug=slug)
    # Get or create ledger
    ledger = LedgerModel.objects.filter(entity=entity).first()
    if not ledger:
        ledger = LedgerModel.objects.create(entity=entity, name="Default Ledger")

    coa = entity.get_default_coa()
    if not coa:
        messages.error(request, "No Chart of Accounts found.")
        return redirect("FixedAssets:asset_dashboard", slug=slug)

    assets = FixedAsset.objects.filter(entity=entity, is_active=True)
    if not assets.exists():
        messages.warning(request, "No active assets to depreciate.")
        return redirect("FixedAssets:dashboard", slug=slug)

    posted = 0
    today = timezone.now()
    month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0) 
    #month_start = today.replace(day=1)

    # Optional: wrap in transaction
    with transaction.atomic():

        assets = FixedAsset.objects.filter(entity=entity, is_active=True)
        print(f"Total assets found: {assets.count()}")  # Debug

        for asset in assets:
            if not asset.category:
                continue

            print(f"Processing asset: {asset.name} (ID: {asset.id})")
            print(f"  - Category: {asset.category}")  # Might be None
            if not asset.category:
                print("  - SKIP: No category")
                continue
            print(f"  - Depreciation method: {asset.category.depreciation_method}")
            print(f"  - Useful life years: {asset.category.useful_life_years}")
            print(
                f"  - Cost: {asset.cost}, Salvage: {asset.salvage_value}, Acc Dep: {asset.accumulated_depreciation}"
            )

            # Check if category has the required GL accounts
            if (
                not asset.category.depreciation_expense_account
                or not asset.category.accumulated_depreciation_account
            ):
                messages.error(
                    request,
                    f"Category '{asset.category.name}' is missing GL accounts for {asset.name}. Please update the category.",
                )
                continue

            # --- Calculate monthly depreciation ---
            if asset.category.depreciation_method == "straight_line":
                useful_months = asset.category.useful_life_years * 12
                monthly_dep = (asset.cost - asset.salvage_value) / Decimal(
                    useful_months
                )
            else:  # reducing balance
                # Rate is in percentage (e.g., 20.00 for 20% per year)
                rate = Decimal(
                    asset.override_depreciation_rate
                    or asset.category.depreciation_rate
                    or 20.00
                )
                monthly_rate = rate / 100 / 12
                nbv = asset.cost - asset.accumulated_depreciation
                monthly_dep = nbv * monthly_rate

            # Prevent over‑depreciation
            if (
                asset.cost - asset.accumulated_depreciation - monthly_dep
            ) < asset.salvage_value:
                monthly_dep = (
                    asset.cost - asset.accumulated_depreciation - asset.salvage_value
                )

            if monthly_dep <= Decimal("0.00"):
                continue

            # --- Get GL accounts ---
            try:
                expense_account = AccountModel.objects.get(
                    coa_model=coa, code=asset.category.depreciation_expense_account.code
                )
                acc_dep_account = AccountModel.objects.get(
                    coa_model=coa,
                    code=asset.category.accumulated_depreciation_account.code,
                )
            except AccountModel.DoesNotExist:
                messages.error(
                    request,
                    f"Account not found for {asset.name}. Check your category accounts.",
                )
                continue

            # --- Create and post journal entry ---
            je = JournalEntryModel.objects.create(
                ledger=ledger,
                timestamp=today,
                description=f"Depreciation for {asset.name}",
                posted=False,
            )

            TransactionModel.objects.create(
                journal_entry=je,
                account=expense_account,
                amount=monthly_dep,
                tx_type="debit",
            )
            TransactionModel.objects.create(
                journal_entry=je,
                account=acc_dep_account,
                amount=monthly_dep,
                tx_type="credit",
            )

            je.posted = True
            je.save()

            # --- Update asset ---
            asset.accumulated_depreciation = (
                asset.accumulated_depreciation or Decimal("0")
            ) + monthly_dep
            asset.book_value = asset.cost - asset.accumulated_depreciation
            asset.last_depreciation_date = today
            asset.save()

            # --- Record depreciation entry ---
            DepreciationEntry.objects.create(
                asset=asset,
                period_start=month_start,
                period_end=today,
                amount=monthly_dep,
                journal_entry=je,
                created_by=request.user,
            )

            posted += 1

    messages.success(request, f"Posted depreciation for {posted} assets.")
    return redirect("FixedAssets:asset_dashboard", slug=slug)


@login_required
def depreciation_schedule(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    assets = FixedAsset.objects.filter(entity=entity, is_active=True).select_related(
        "category"
    )

    asset_data = []
    for asset in assets:
        total_dep = DepreciationEntry.objects.filter(asset=asset).aggregate(
            Sum("amount")
        )["amount__sum"] or Decimal("0")
        book_value = asset.cost - total_dep
        asset_data.append(
            {
                "name": asset.name,
                "cost": asset.cost,
                "total_depreciation": total_dep,
                "book_value": book_value,
                "depreciation_entries": DepreciationEntry.objects.filter(
                    asset=asset
                ).order_by("-period_end")[:12],
            }
        )

    context = {
        "entity": entity,
        "asset_data": asset_data,
        "today": timezone.now().date(),
    }
    return render(request, "FixedAssets/depreciation_schedule.html", context)


@login_required
def depreciation_list_manage(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    # Get all depreciation entries with related asset
    entries = DepreciationEntry.objects.select_related('asset', 'journal_entry').all().order_by('-period_end')
    print(entries)
    
    # Calculate totals
    total_depreciation = entries.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Group by asset (optional summary)
    asset_summary = {}
    for entry in entries:
        asset_name = entry.asset.name
        if asset_name not in asset_summary:
            asset_summary[asset_name] = 0
        asset_summary[asset_name] += entry.amount
    
    context = {
        'entries': entries,
        'total_depreciation': total_depreciation,
        'asset_summary': asset_summary,
        'entry_count': entries.count(),
    }
    return render(request, 'FixedAssets/depreciation_list_manage.html', context)


@login_required
def fixed_asset_register(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    assets = FixedAsset.objects.filter(is_active=True).select_related('category')
    data = []
    total_cost = Decimal('0')
    total_acc_dep = Decimal('0')
    total_nbv = Decimal('0')
    for asset in assets:
        acc_dep = asset.depreciation_entries.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        nbv = asset.cost - acc_dep
        data.append({
            'asset': asset,
            'acc_dep': acc_dep,
            'nbv': nbv,
        })
        total_cost += asset.cost
        total_acc_dep += acc_dep
        total_nbv += nbv
    context = {
        'data': data,
        'total_cost': total_cost,
        'total_acc_dep': total_acc_dep,
        'total_nbv': total_nbv,
    }
    return render(request, 'FixedAssets/fixed_assets_register.html', context)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import DepreciationEntry

@login_required
def depreciation_export_excel(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    wb = Workbook()
    ws = wb.active
    ws.title = "Depreciation Schedule"

    # --- Headers ---
    headers = ['Asset ID', 'Asset Name', 'Period Start', 'Period End', 'Amount', 'Created At']
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Data ---
    entries = DepreciationEntry.objects.select_related('asset').all().order_by('-period_end')
    
    for row_idx, entry in enumerate(entries, 2):
        ws.cell(row=row_idx, column=1, value=entry.asset_id)
        ws.cell(row=row_idx, column=2, value=entry.asset.name)
        ws.cell(row=row_idx, column=3, value=entry.period_start.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=4, value=entry.period_end.strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=5, value=float(entry.amount))
        ws.cell(row=row_idx, column=6, value=entry.created_at.strftime("%Y-%m-%d %H:%M"))

    # --- Format numbers ---
    for row in ws.iter_rows(min_row=2, max_row=len(entries)+1, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'

    # --- Total row ---
    total_row = len(entries) + 2
    ws.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=5, value=sum(e.amount for e in entries))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'

    # --- Auto-width ---
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # --- Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=depreciation_schedule.xlsx'
    wb.save(response)
    return response


from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from .models import DepreciationEntry

@login_required
def depreciation_export_pdf(request, slug):
    from django_ledger.models import EntityModel
    entity = get_object_or_404(EntityModel, slug=slug)
    entries = DepreciationEntry.objects.select_related('asset').all().order_by('-period_end')
    total_depreciation = sum(e.amount for e in entries)

    template = get_template('FixedAssets/depreciation_pdf.html')
    html = template.render({
        'entries': entries,
        'total_depreciation': total_depreciation,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=depreciation_schedule.pdf'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF: ' + str(pisa_status.err))
    return response


# views.py
def category_assets_list(request, slug, category_id):
    entity = get_object_or_404(EntityModel, slug=slug)
    category = get_object_or_404(AssetCategory, id=category_id, entity=entity)
    assets = FixedAsset.objects.filter(entity=entity, category=category, is_active=True)
    return render(
        request,
        "FixedAssets/category_assets.html",
        {"category": category, "assets": assets},
    )


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from decimal import Decimal
from django.utils import timezone
from datetime import timedelta
from django_ledger.models import EntityModel
from .models import FixedAsset, DepreciationEntry


@login_required
def fixed_assets_register(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)

    # Get all active assets for this entity
    assets = FixedAsset.objects.filter(entity=entity, is_active=True).select_related(
        "category"
    )

    registrar_data = []
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)

    for asset in assets:
        # Get total accumulated depreciation
        total_dep = DepreciationEntry.objects.filter(asset=asset).aggregate(
            Sum("amount")
        )["amount__sum"] or Decimal("0.00")

        # Get depreciation for the current year (year-to-date)
        year_dep = DepreciationEntry.objects.filter(
            asset=asset, period_start__gte=year_start, period_end__lte=today
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")

        # Calculate current value (cost - total accumulated depreciation)
        current_value = asset.cost - total_dep

        # Impairment = cost - accumulated depreciation - salvage value (if negative, set to 0)
        impairment = current_value - asset.salvage_value
        if impairment < 0:
            impairment = Decimal("0.00")

        registrar_data.append(
            {
                "asset": asset,
                "category": asset.category,
                "cost": asset.cost,
                "salvage_value": asset.salvage_value,
                "total_accumulated_depreciation": total_dep,
                "year_depreciation": year_dep,
                "current_value": current_value,
                "impairment": impairment,
                "depreciation_rate": (
                    asset.category.depreciation_rate if asset.category else None
                ),
                "asset_account": (
                    asset.category.asset_account if asset.category else None
                ),
                "accumulated_depreciation_account": (
                    asset.category.accumulated_depreciation_account
                    if asset.category
                    else None
                ),
                "depreciation_expense_account": (
                    asset.category.depreciation_expense_account
                    if asset.category
                    else None
                ),
            }
        )

    context = {
        "entity": entity,
        "registrar_data": registrar_data,
        "today": today,
    }
    return render(request, "FixedAssets/fixed_assets_register.html", context)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django.shortcuts import get_object_or_404
from django_ledger.models import EntityModel
from django.db.models import Sum
from .models import FixedAsset, DepreciationEntry
from django.utils import timezone


def fixed_assets_register_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    assets = FixedAsset.objects.filter(entity=entity, is_active=True).select_related(
        "category"
    )
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fixed Assets Registrar"

    # --- Headers ---
    headers = [
        "#",
        "Category",
        "Asset",
        "Depreciation Rate (%)",
        "Cost",
        "Salvage Value",
        "Asset Account",
        "Accumulated Depreciation Account",
        "Depreciation Expense Account",
        "Accumulated Depreciation",
        "Current Value",
        "Year Depreciation",
        "Impairment",
    ]
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Data ---
    row = 2
    total_cost = Decimal("0")
    total_salvage = Decimal("0")
    total_acc_dep = Decimal("0")
    total_current = Decimal("0")
    total_year_dep = Decimal("0")
    total_impairment = Decimal("0")

    for idx, asset in enumerate(assets, 1):
        category = asset.category
        total_dep = DepreciationEntry.objects.filter(asset=asset).aggregate(
            Sum("amount")
        )["amount__sum"] or Decimal("0")
        year_dep = DepreciationEntry.objects.filter(
            asset=asset, period_start__gte=year_start, period_end__lte=today
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0")

        current_value = asset.cost - total_dep
        impairment = current_value - asset.salvage_value
        if impairment < 0:
            impairment = Decimal("0")

        total_cost += asset.cost
        total_salvage += asset.salvage_value
        total_acc_dep += total_dep
        total_current += current_value
        total_year_dep += year_dep
        total_impairment += impairment

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=category.name if category else "—")
        ws.cell(row=row, column=3, value=asset.name)
        ws.cell(
            row=row,
            column=4,
            value=(
                float(category.depreciation_rate)
                if category and category.depreciation_rate
                else None
            ),
        )
        ws.cell(row=row, column=5, value=float(asset.cost))
        ws.cell(row=row, column=6, value=float(asset.salvage_value))
        ws.cell(
            row=row,
            column=7,
            value=(
                category.asset_account.code
                if category and category.asset_account
                else "—"
            ),
        )
        ws.cell(
            row=row,
            column=8,
            value=(
                category.accumulated_depreciation_account.code
                if category and category.accumulated_depreciation_account
                else "—"
            ),
        )
        ws.cell(
            row=row,
            column=9,
            value=(
                category.depreciation_expense_account.code
                if category and category.depreciation_expense_account
                else "—"
            ),
        )
        ws.cell(row=row, column=10, value=float(total_dep))
        ws.cell(row=row, column=11, value=float(current_value))
        ws.cell(row=row, column=12, value=float(year_dep))
        ws.cell(row=row, column=13, value=float(impairment))

        row += 1

    # --- Totals Row ---
    ws.cell(row=row, column=4, value="TOTALS").font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_cost)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_salvage)).font = Font(bold=True)
    ws.cell(row=row, column=10, value=float(total_acc_dep)).font = Font(bold=True)
    ws.cell(row=row, column=11, value=float(total_current)).font = Font(bold=True)
    ws.cell(row=row, column=12, value=float(total_year_dep)).font = Font(bold=True)
    ws.cell(row=row, column=13, value=float(total_impairment)).font = Font(bold=True)

    # --- Auto-width ---
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Response ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Fixed_Assets_Register_{entity.slug}_{today}.xlsx"'
    )
    wb.save(response)
    return response


from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.shortcuts import get_object_or_404
from django_ledger.models import EntityModel
from django.db.models import Sum
from .models import FixedAsset, DepreciationEntry
from decimal import Decimal
from django.utils import timezone


def fixed_assets_register_PDF(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    assets = FixedAsset.objects.filter(entity=entity, is_active=True).select_related(
        "category"
    )
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)

    registrar_data = []
    total_cost = Decimal("0")
    total_salvage = Decimal("0")
    total_acc_dep = Decimal("0")
    total_current = Decimal("0")
    total_year_dep = Decimal("0")
    total_impairment = Decimal("0")

    for asset in assets:
        category = asset.category
        total_dep = DepreciationEntry.objects.filter(asset=asset).aggregate(
            Sum("amount")
        )["amount__sum"] or Decimal("0")
        year_dep = DepreciationEntry.objects.filter(
            asset=asset, period_start__gte=year_start, period_end__lte=today
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0")
        current_value = asset.cost - total_dep
        impairment = current_value - asset.salvage_value
        if impairment < 0:
            impairment = Decimal("0")

        total_cost += asset.cost
        total_salvage += asset.salvage_value
        total_acc_dep += total_dep
        total_current += current_value
        total_year_dep += year_dep
        total_impairment += impairment

        registrar_data.append(
            {
                "asset": asset,
                "category": category,
                "cost": asset.cost,
                "salvage_value": asset.salvage_value,
                "total_accumulated_depreciation": total_dep,
                "year_depreciation": year_dep,
                "current_value": current_value,
                "impairment": impairment,
                "depreciation_rate": category.depreciation_rate if category else None,
                "asset_account": category.asset_account if category else None,
                "accumulated_depreciation_account": (
                    category.accumulated_depreciation_account if category else None
                ),
                "depreciation_expense_account": (
                    category.depreciation_expense_account if category else None
                ),
            }
        )

    context = {
        "entity": entity,
        "registrar_data": registrar_data,
        "total_cost": total_cost,
        "total_salvage": total_salvage,
        "total_acc_dep": total_acc_dep,
        "total_current_value": total_current,
        "total_year_dep": total_year_dep,
        "total_impairment": total_impairment,
        "today": today,
    }

    template = get_template("FixedAssets/fixed_assets_register_pdf.html")
    html = template.render(context)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="Fixed_Assets_Register_{entity.slug}_{today}.pdf"'
    )
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF")
    return response


@login_required
def fixed_assets_register_list(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)

    # Get all active assets for this entity
    assets = FixedAsset.objects.filter(entity=entity, is_active=True).select_related(
        "category"
    )

    registrar_data = []
    today = timezone.now().date()
    year_start = today.replace(month=1, day=1)

    for asset in assets:
        # Get total accumulated depreciation
        total_dep = DepreciationEntry.objects.filter(asset=asset).aggregate(
            Sum("amount")
        )["amount__sum"] or Decimal("0.00")

        # Get depreciation for the current year (year-to-date)
        year_dep = DepreciationEntry.objects.filter(
            asset=asset, period_start__gte=year_start, period_end__lte=today
        ).aggregate(Sum("amount"))["amount__sum"] or Decimal("0.00")

        # Calculate current value (cost - total accumulated depreciation)
        current_value = asset.cost - total_dep

        # Impairment = cost - accumulated depreciation - salvage value (if negative, set to 0)
        impairment = current_value - asset.salvage_value
        if impairment < 0:
            impairment = Decimal("0.00")

        registrar_data.append(
            {
                "asset": asset,
                "category": asset.category,
                "cost": asset.cost,
                "salvage_value": asset.salvage_value,
                "total_accumulated_depreciation": total_dep,
                "year_depreciation": year_dep,
                "current_value": current_value,
                "impairment": impairment,
                "depreciation_rate": (
                    asset.category.depreciation_rate if asset.category else None
                ),
                "asset_account": (
                    asset.category.asset_account if asset.category else None
                ),
                "accumulated_depreciation_account": (
                    asset.category.accumulated_depreciation_account
                    if asset.category
                    else None
                ),
                "depreciation_expense_account": (
                    asset.category.depreciation_expense_account
                    if asset.category
                    else None
                ),
            }
        )

    context = {
        "entity": entity,
        "registrar_data": registrar_data,
        "today": today,
    }
    return render(request, "FixedAssets/fixed_assets_register.html", context)
