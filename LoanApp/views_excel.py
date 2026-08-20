# views.py - Add this function
import openpyxl
from openpyxl.styles import Font, Alignment
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from decimal import Decimal
import json
from datetime import date, datetime, timedelta
from django.contrib.auth.decorators import login_required

from django.http import HttpResponse


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from io import BytesIO
from reportlab.rl_settings import underlineWidth
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ## Tables
from .models import Loan, Guarantor
from MembersApp.models import Master
from django_ledger.models import EntityModel

def gua_list_excel(request, slug):
    entity=get_object_or_404(EntityModel, slug=slug)
    """Export loans list as Excel"""
    # Get loans data
    loans = Loan.objects.all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan List"
    
    # Add headers
    headers = ['Loan ID', 'Member ID', 'Member Name', 'Principal', 'Interest Rate', 
               'Term', 'Shortfall', 'Guarantor Count', 'Guarantee Total', 'Guarantor Details']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data rows
    for row, loan in enumerate(loans, 2):
        # Get guarantor details
        guarantor_details = []
        if loan.guarantor_data and 'guarantors' in loan.guarantor_data:
            for g in loan.guarantor_data['guarantors']:
                guarantor_details.append(f"{g['name']}: GHS {g['amount']} on {g['date']}")
        
        ws.cell(row=row, column=1, value=loan.id)
        ws.cell(row=row, column=2, value=loan.master_id)
        ws.cell(row=row, column=3, value=loan.master_name)
        ws.cell(row=row, column=4, value=float(loan.principal))
        ws.cell(row=row, column=5, value=float(loan.interest_rate))
        ws.cell(row=row, column=6, value=loan.loan_term)
        ws.cell(row=row, column=7, value=float(loan.shortfall))
        ws.cell(row=row, column=8, value=loan.guarantor_count)
        ws.cell(row=row, column=9, value=float(loan.total_guaranteed))
        ws.cell(row=row, column=10, value="; ".join(guarantor_details) if guarantor_details else "None")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="loan_list.xlsx"'
    
    wb.save(response)
    return response


@login_required
def loan_list_financials_excel(request, slug):
    entity=get_object_or_404(EntityModel, slug=slug)
    """Generate Excel file of all loans"""
    loans = Loan.objects.select_related('master').all()

    # Prepare data for DataFrame
    data = []
    for loan in loans:
        data.append({
            'Loan ID': loan.id,
            'Borrower': loan.master.full_name,
            'Principal': float(loan.principal) if loan.principal else 0,
            'Interest Rate (%)': float(loan.interest_rate) if loan.interest_rate else 0,
            'Term (Months)': loan.loan_term,
            'Total Interest': float(loan.tot_int) if loan.tot_int else 0,
            'Total Repayable': float(loan.tot_ded) if loan.tot_ded else 0,
            'Loan Balance': float(loan.loan_balance) if loan.loan_balance else 0,
            'Months Remaining': loan.months_remain,
            'Due Interest': float(loan.due_interest) if loan.due_interest else 0,
            'Due Repayment': float(loan.due_repayment) if loan.due_repayment else 0,
            'Monthly Repayment': float(loan.monthly_repayment) if loan.monthly_repayment else 0,
            'Status': loan.payment_status,
        })

    df = pd.DataFrame(data)

    # Create Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="loans_report.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Loans', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Loans']

        # Add totals row
        totals_row = len(data) + 2
        worksheet[f'A{totals_row}'] = 'TOTALS'
        worksheet[f'C{totals_row}'] = f"=SUM(C2:C{len(data)+1})"
        worksheet[f'E{totals_row}'] = f"=SUM(E2:E{len(data)+1})"
        worksheet[f'F{totals_row}'] = f"=SUM(F2:F{len(data)+1})"
        worksheet[f'G{totals_row}'] = f"=SUM(G2:G{len(data)+1})"
        worksheet[f'H{totals_row}'] = f"=SUM(H2:H{len(data)+1})"

        # Format currency columns
        for col in ['C', 'F', 'G', 'H', 'J', 'K', 'L']:
            for row in range(2, len(data) + 2):
                cell = worksheet[f'{col}{row}']
                cell.number_format = '#,##0.00'

    return response


def loan_update_list_excel(request, slug):
    entity = get_object_or_404(EntityModel, slug=slug)
    loans = Loan.objects.filter(entity=entity).select_related("master").order_by("-id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan List"

    headers = [
        "ID",
        "Member",
        "Principal",
        "Interest %",
        "Term",
        "Total Interest",
        "Total Deductions",
        "Balance",
        "Last Interest Paid",
        "Last Repayment",
        "Last Payment Date",
        "Due Days",
        "Due Interest",
        "Due Repayment",
        "Due Total",
        "Disbursement Date",
        "Due Date",
        "Overdue Days",
        "Status",
        "Loan Credit Balance",
        "Next Repayment Date",
        "Expiry Date",
        "Update Count",
    ]

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, loan in enumerate(loans, 2):
        ws.cell(row=row_idx, column=1, value=loan.id)
        ws.cell(row=row_idx, column=2, value=loan.master.full_name)
        ws.cell(row=row_idx, column=3, value=float(loan.principal))
        ws.cell(row=row_idx, column=4, value=float(loan.interest_rate))
        ws.cell(row=row_idx, column=5, value=loan.loan_term)
        ws.cell(row=row_idx, column=6, value=float(loan.tot_int) if loan.tot_int else 0)
        ws.cell(row=row_idx, column=7, value=float(loan.tot_ded) if loan.tot_ded else 0)
        ws.cell(row=row_idx, column=8, value=float(loan.loan_balance))
        ws.cell(
            row=row_idx,
            column=9,
            value=float(loan.last_interest_paid) if loan.last_interest_paid else 0,
        )
        ws.cell(
            row=row_idx,
            column=10,
            value=float(loan.last_repayment_paid) if loan.last_repayment_paid else 0,
        )
        ws.cell(
            row=row_idx,
            column=11,
            value=(
                loan.last_payment_date.strftime("%d/%m/%Y")
                if loan.last_payment_date
                else ""
            ),
        )
        ws.cell(row=row_idx, column=12, value=loan.due_days if loan.due_days else 0)
        ws.cell(
            row=row_idx,
            column=13,
            value=float(loan.due_interest) if loan.due_interest else 0,
        )
        ws.cell(
            row=row_idx,
            column=14,
            value=float(loan.due_repayment) if loan.due_repayment else 0,
        )
        ws.cell(
            row=row_idx,
            column=15,
            value=float(loan.due_tot_repayment) if loan.due_tot_repayment else 0,
        )
        ws.cell(
            row=row_idx,
            column=16,
            value=(
                loan.disbursement_date.strftime("%d/%m/%Y")
                if loan.disbursement_date
                else ""
            ),
        )
        ws.cell(
            row=row_idx,
            column=17,
            value=loan.due_date.strftime("%d/%m/%Y") if loan.due_date else "",
        )
        ws.cell(
            row=row_idx, column=18, value=loan.overdue_days if loan.overdue_days else 0
        )
        ws.cell(row=row_idx, column=19, value=loan.status if loan.status else "")
        ws.cell(
            row=row_idx,
            column=20,
            value=float(loan.loan_credit_balance) if loan.loan_credit_balance else 0,
        )
        ws.cell(
            row=row_idx,
            column=21,
            value=(
                loan.next_repayment_date.strftime("%d/%m/%Y")
                if loan.next_repayment_date
                else ""
            ),
        )
        ws.cell(
            row=row_idx,
            column=22,
            value=loan.expiry_date.strftime("%d/%m/%Y") if loan.expiry_date else "",
        )
        ws.cell(
            row=row_idx,
            column=23,
            value=loan.loan_update_cnt if loan.loan_update_cnt else 0,
        )

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 16

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Loans_{entity.slug}_{datetime.now().date()}.xlsx"'
    )
    wb.save(response)
    return response
